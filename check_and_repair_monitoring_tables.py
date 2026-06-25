#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查并修复小红书/抖音监控总表与 origin_data 的核心字段。

这个脚本默认不重新访问平台，优先使用三类本地数据互相补齐：
1. Pipeline/*_origin_data.csv 中的原始抓取字段
2. Pipeline/*_Data_Table_on_Channel_Public_Opinion_Monitoring_2026.csv 中的监控字段
3. Comment_Data/{xhs,dy} 下已经导出的逐帖评论文件

如果监控表和 origin_data 对同一条数据都缺少核心字段，可加 --refresh-missing
按缺失帖子逐条回源重爬。该模式会使用可见浏览器、默认登录态、慢速间隔和本地
历史 xsec_token 链接缓存，尽量减少触发平台风控。

用法：
  python3 check_and_repair_monitoring_tables.py
  python3 check_and_repair_monitoring_tables.py --platform xhs --dry-run
  python3 check_and_repair_monitoring_tables.py --recompute-interactions
  python3 check_and_repair_monitoring_tables.py --platform all --refresh-missing --max-refresh 20 --headed
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import random
import re
import shutil
import subprocess
import sys
import time
import zipfile
from collections import OrderedDict, defaultdict
from xml.etree import ElementTree as ET
from pathlib import Path
from typing import Any, Callable


PROJECT_ROOT = Path(__file__).resolve().parent
PIPELINE_DIR = PROJECT_ROOT / "Pipeline"
COMMENT_DIR = PROJECT_ROOT / "Comment_Data"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

import dy_common  # type: ignore  # noqa: E402
import xhs_note_to_csv  # type: ignore  # noqa: E402
from pipeline_paths import (  # type: ignore  # noqa: E402
    DY_DATA_TABLE_CSV,
    DY_ORIGIN_CSV,
    XHS_DATA_TABLE_CSV,
    XHS_ORIGIN_CSV,
)


CORE_FIELDS = [
    "发布时间",
    "笔记标题",
    "笔记链接",
    "笔记内容",
    "点赞量",
    "收藏量",
    "评论量",
    "分享量",
    "互动量",
    "博主昵称",
    "渠道类型",
    "笔记ID",
]

ORIGIN_REPAIR_FIELDS = [
    "repair.笔记ID",
    "repair.博主昵称",
    "repair.笔记链接",
    "repair.笔记标题",
    "repair.笔记内容",
    "repair.点赞量",
    "repair.收藏量",
    "repair.评论量",
    "repair.分享量",
    "repair.互动量",
    "repair.发布时间",
    "repair.渠道类型",
    "repair.comment_rows",
    "repair.comment_file",
    "repair.updated_at",
]

XHS_URL_RE = re.compile(
    r"https?://(?:www\.)?xiaohongshu\.com/(?:discovery/item|explore|search_result)/[^\s，。！？,，）)】\]]+"
)
XHS_ID_RE = re.compile(r"(?:/explore/|/discovery/item/|/search_result/)([0-9a-zA-Z]{24})")
DY_URL_RE = re.compile(
    r"https?://(?:(?:www|v)\.)?(?:douyin\.com|iesdouyin\.com)/[^\s，。！？,，）)】\]]+"
)
DY_ID_RE = re.compile(r"(?:modal_id=|/video/|/note/|/share/video/|/discover/)(\d{10,30})")

PLATFORMS = {
    "xhs": {
        "name": "小红书",
        "channel": "小红书",
        "origin": XHS_ORIGIN_CSV,
        "table": XHS_DATA_TABLE_CSV,
        "summary": xhs_note_to_csv.xhs_summary_row,
        "url_re": XHS_URL_RE,
        "id_re": XHS_ID_RE,
    },
    "dy": {
        "name": "抖音",
        "channel": "抖音",
        "origin": DY_ORIGIN_CSV,
        "table": DY_DATA_TABLE_CSV,
        "summary": dy_common.summary_row,
        "url_re": DY_URL_RE,
        "id_re": DY_ID_RE,
    },
}

NOTE_SCRIPTS = {
    "xhs": PIPELINE_DIR / "xhs_note_to_csv.py",
    "dy": PIPELINE_DIR / "dy_note_to_csv.py",
}

REFRESH_TMP_DIR = PIPELINE_DIR / "gui_exports" / "repair_refresh"

DEFAULT_MISSING_FIELDS = [
    "笔记内容",
    "笔记标题",
    "博主昵称",
    "发布时间",
    "点赞量",
    "收藏量",
    "评论量",
    "分享量",
]


def python_can_run(python_bin: str) -> bool:
    try:
        result = subprocess.run(
            [python_bin, "-c", "import sys; print(sys.executable)"],
            cwd=str(PROJECT_ROOT),
            text=True,
            capture_output=True,
            check=False,
            timeout=5,
        )
        return result.returncode == 0
    except Exception:
        return False


def choose_script_python() -> str:
    raw = [
        os.environ.get("PIPELINE_PYTHON"),
        str(PROJECT_ROOT / ".venv" / "bin" / "python"),
        "/Library/Developer/CommandLineTools/usr/bin/python3",
        "/usr/bin/python3",
        shutil.which("python3.9"),
        shutil.which("python3"),
        sys.executable,
        "python3",
    ]
    seen: set[str] = set()
    for item in raw:
        if not item:
            continue
        candidate = shutil.which(item) if "/" not in item else item
        candidate = candidate or item
        if candidate in seen:
            continue
        seen.add(candidate)
        if "/" in candidate and not Path(candidate).exists():
            continue
        if python_can_run(candidate):
            return candidate
    return sys.executable or "python3"


SCRIPT_PYTHON = choose_script_python()


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def is_blank(value: Any) -> bool:
    text = clean_text(value)
    return text == "" or text.lower() in {"none", "nan", "null", "undefined"}


def number_value(value: Any) -> int:
    text = clean_text(value).replace(",", "")
    if not text:
        return 0
    for suffix, factor in (("万", 10000), ("千", 1000), ("w", 10000), ("W", 10000), ("k", 1000), ("K", 1000)):
        if text.endswith(suffix):
            try:
                return int(float(text[: -len(suffix)]) * factor)
            except Exception:
                return 0
    try:
        return int(float(text))
    except Exception:
        return 0


def interaction_total(row: dict[str, Any]) -> str:
    total = sum(number_value(row.get(field)) for field in ("点赞量", "收藏量", "评论量", "分享量"))
    return str(total)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists() or path.stat().st_size == 0:
        return [], []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fields: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def ensure_fields(fields: list[str], rows: list[dict[str, Any]], required: list[str]) -> tuple[list[str], bool]:
    out = list(fields)
    changed = False
    for field in required:
        if field not in out:
            out.append(field)
            changed = True
    for row in rows:
        for field in out:
            row.setdefault(field, "")
    return out, changed


def decode_urlish(value: str) -> str:
    return html.unescape(str(value or "")).replace("\\u0026", "&").replace("\\/", "/")


def extract_platform_url(platform: str, value: Any) -> str:
    config = PLATFORMS[platform]
    text = decode_urlish(str(value or ""))
    match = config["url_re"].search(text)
    url = match.group(0) if match else (text if text.startswith("http") else "")
    return url.rstrip("。；;，,）)】]")


def extract_note_id(platform: str, value: Any) -> str:
    text = decode_urlish(str(value or ""))
    if platform == "dy":
        url = extract_platform_url(platform, text) or text
        note_id = dy_common.parse_aweme_id_from_url(url)
        if note_id:
            return note_id
    match = PLATFORMS[platform]["id_re"].search(text)
    return match.group(1) if match else ""


def canonical_link(platform: str, note_id: str, current: Any = "") -> str:
    url = extract_platform_url(platform, current)
    if url:
        return url
    if not note_id:
        return clean_text(current)
    if platform == "dy":
        return dy_common.canonical_aweme_url(note_id)
    return f"https://www.xiaohongshu.com/discovery/item/{note_id}"


def first_nonblank(row: dict[str, Any], names: list[str]) -> str:
    for name in names:
        value = row.get(name)
        if not is_blank(value):
            return clean_text(value)
    return ""


def fallback_summary_from_row(platform: str, row: dict[str, Any]) -> dict[str, Any]:
    note_id = first_nonblank(row, ["笔记ID", "repair.笔记ID", "note_id", "aweme_id", "aweme_detail.aweme_id"])
    link = first_nonblank(row, ["笔记链接", "repair.笔记链接", "source_url", "aweme_detail.share_url"])
    note_id = note_id or extract_note_id(platform, link) or extract_note_id(platform, json.dumps(row, ensure_ascii=False)[:5000])
    link = canonical_link(platform, note_id, link)
    return {
        "笔记ID": note_id,
        "博主昵称": first_nonblank(row, ["博主昵称", "repair.博主昵称", "aweme_detail.author.nickname"]),
        "笔记链接": link,
        "笔记标题": first_nonblank(row, ["笔记标题", "repair.笔记标题", "title", "display_title"]),
        "笔记内容": first_nonblank(row, ["笔记内容", "repair.笔记内容", "desc", "content", "aweme_detail.desc"]),
        "点赞量": first_nonblank(row, ["点赞量", "repair.点赞量", "aweme_detail.statistics.digg_count"]),
        "收藏量": first_nonblank(row, ["收藏量", "repair.收藏量", "aweme_detail.statistics.collect_count"]),
        "评论量": first_nonblank(row, ["评论量", "repair.评论量", "aweme_detail.statistics.comment_count"]),
        "分享量": first_nonblank(row, ["分享量", "repair.分享量", "aweme_detail.statistics.share_count"]),
        "发布时间": first_nonblank(row, ["发布时间", "repair.发布时间", "aweme_detail.create_time"]),
        "渠道类型": first_nonblank(row, ["渠道类型", "repair.渠道类型"]) or PLATFORMS[platform]["channel"],
    }


def merge_summary(primary: dict[str, Any], fallback: dict[str, Any]) -> dict[str, Any]:
    merged = dict(primary)
    for field in CORE_FIELDS:
        if field == "互动量":
            continue
        if is_blank(merged.get(field)) and not is_blank(fallback.get(field)):
            merged[field] = fallback[field]
    if not is_blank(merged.get("笔记链接")):
        pass
    elif not is_blank(fallback.get("笔记链接")):
        merged["笔记链接"] = fallback["笔记链接"]
    merged["互动量"] = interaction_total(merged)
    return merged


def summarize_origin_row(
    platform: str,
    origin: dict[str, Any],
    summary_func: Callable[[dict[str, Any]], OrderedDict],
) -> dict[str, Any]:
    try:
        summary = dict(summary_func(origin))
    except Exception:
        summary = {}
    summary = merge_summary(summary, fallback_summary_from_row(platform, origin))
    note_id = clean_text(summary.get("笔记ID")) or extract_note_id(platform, summary.get("笔记链接"))
    summary["笔记ID"] = note_id
    summary["笔记链接"] = canonical_link(platform, note_id, summary.get("笔记链接"))
    summary["渠道类型"] = clean_text(summary.get("渠道类型")) or PLATFORMS[platform]["channel"]
    summary["互动量"] = interaction_total(summary)
    return summary


def candidate_score(row: dict[str, Any]) -> int:
    body = clean_text(row.get("笔记内容"))
    title = clean_text(row.get("笔记标题"))
    score = min(len(body), 1200) * 3 + min(len(title), 200) * 2
    for field in ("发布时间", "点赞量", "收藏量", "评论量", "分享量", "笔记链接", "博主昵称"):
        if not is_blank(row.get(field)):
            score += 30
    if body and body != title:
        score += 150
    return score


def comment_count_from_xlsx(path: Path) -> tuple[int, set[str]]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return comment_count_from_xlsx_zip(path)
    ids: set[str] = set()
    count = 0
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        headers = [clean_text(cell) for cell in next(iterator, [])]
        id_index = headers.index("笔记ID") if "笔记ID" in headers else -1
        content_index = headers.index("评论内容") if "评论内容" in headers else -1
        for values in iterator:
            if not values:
                continue
            content = clean_text(values[content_index]) if content_index >= 0 and content_index < len(values) else ""
            if not content and not any(clean_text(item) for item in values):
                continue
            count += 1
            if id_index >= 0 and id_index < len(values):
                note_id = clean_text(values[id_index])
                if note_id:
                    ids.add(note_id)
        wb.close()
    except Exception:
        return comment_count_from_xlsx_zip(path)
    return count, ids


def column_index(cell_ref: str) -> int:
    letters = re.match(r"([A-Z]+)", cell_ref or "")
    if not letters:
        return 0
    value = 0
    for char in letters.group(1):
        value = value * 26 + (ord(char) - ord("A") + 1)
    return max(value - 1, 0)


def read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except Exception:
        return []
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    values: list[str] = []
    for item in root.findall(f".//{ns}si"):
        parts = [node.text or "" for node in item.findall(f".//{ns}t")]
        values.append("".join(parts))
    return values


def cell_value(cell: ET.Element, shared: list[str]) -> str:
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(f".//{ns}t"))
    value_node = cell.find(f"{ns}v")
    raw = value_node.text if value_node is not None else ""
    if cell_type == "s":
        try:
            return shared[int(raw)]
        except Exception:
            return ""
    return raw or ""


def first_sheet_rows_zip(path: Path) -> list[list[str]]:
    try:
        with zipfile.ZipFile(path) as zf:
            sheet_names = sorted(name for name in zf.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name))
            if not sheet_names:
                return []
            shared = read_shared_strings(zf)
            root = ET.fromstring(zf.read(sheet_names[0]))
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            rows: list[list[str]] = []
            for row in root.findall(f".//{ns}sheetData/{ns}row"):
                values: list[str] = []
                for cell in row.findall(f"{ns}c"):
                    index = column_index(cell.attrib.get("r", ""))
                    while len(values) <= index:
                        values.append("")
                    values[index] = clean_text(cell_value(cell, shared))
                if any(values):
                    rows.append(values)
            return rows
    except Exception:
        return []


def comment_count_from_xlsx_zip(path: Path) -> tuple[int, set[str]]:
    rows = first_sheet_rows_zip(path)
    if not rows:
        return 0, set()
    headers = rows[0]
    id_index = headers.index("笔记ID") if "笔记ID" in headers else -1
    content_index = headers.index("评论内容") if "评论内容" in headers else -1
    count = 0
    ids: set[str] = set()
    for values in rows[1:]:
        content = values[content_index] if content_index >= 0 and content_index < len(values) else ""
        if not content and not any(values):
            continue
        count += 1
        if id_index >= 0 and id_index < len(values) and values[id_index]:
            ids.add(values[id_index])
    return count, ids


def comment_count_from_csv(path: Path) -> tuple[int, set[str]]:
    ids: set[str] = set()
    try:
        fields, rows = read_csv(path)
    except Exception:
        return 0, set()
    count = 0
    for row in rows:
        if is_blank(row.get("评论内容")) and not any(clean_text(value) for value in row.values()):
            continue
        count += 1
        note_id = clean_text(row.get("笔记ID"))
        if note_id:
            ids.add(note_id)
    return count, ids


def build_comment_index(platform: str) -> dict[str, dict[str, Any]]:
    directory = COMMENT_DIR / platform
    index: dict[str, dict[str, Any]] = defaultdict(lambda: {"commentRows": 0, "files": []})
    if not directory.exists():
        return {}
    for path in sorted(list(directory.glob("*.xlsx")) + list(directory.glob("*.csv"))):
        if path.suffix.lower() == ".xlsx":
            count, ids = comment_count_from_xlsx(path)
        else:
            count, ids = comment_count_from_csv(path)
        if count <= 0:
            continue
        keys = set(ids) or {path.stem}
        for note_id in keys:
            item = index[note_id]
            item["commentRows"] = max(int(item.get("commentRows") or 0), count)
            item["files"].append(str(path.relative_to(PROJECT_ROOT)))
    return dict(index)


def update_origin_row(
    platform: str,
    origin: dict[str, Any],
    summary: dict[str, Any],
    comment_index: dict[str, dict[str, Any]],
) -> tuple[bool, list[str]]:
    changed = False
    changed_fields: list[str] = []
    note_id = clean_text(summary.get("笔记ID"))
    if not note_id:
        return False, []

    for raw_id_field in (["note_id", "aweme_id"] if platform == "dy" else ["note_id"]):
        if is_blank(origin.get(raw_id_field)):
            origin[raw_id_field] = note_id
            changed = True
            changed_fields.append(raw_id_field)
    if is_blank(origin.get("source_url")):
        origin["source_url"] = canonical_link(platform, note_id, summary.get("笔记链接"))
        changed = True
        changed_fields.append("source_url")

    for field in CORE_FIELDS:
        if field == "互动量":
            value = interaction_total(summary)
        else:
            value = summary.get(field, "")
        repair_field = f"repair.{field}"
        old = clean_text(origin.get(repair_field))
        new = clean_text(value)
        if new and (not old or (field == "笔记内容" and len(new) > len(old))):
            origin[repair_field] = new
            changed = True
            changed_fields.append(repair_field)

    comment = comment_index.get(note_id)
    if comment:
        rows = str(comment.get("commentRows") or "")
        files = " | ".join(comment.get("files") or [])
        for field, value in (("repair.comment_rows", rows), ("repair.comment_file", files)):
            if value and clean_text(origin.get(field)) != value:
                origin[field] = value
                changed = True
                changed_fields.append(field)
    if changed:
        origin["repair.updated_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    return changed, changed_fields


def build_minimal_origin_row(platform: str, fields: list[str], summary: dict[str, Any], comment_index: dict[str, dict[str, Any]]) -> dict[str, Any]:
    row = {field: "" for field in fields}
    note_id = clean_text(summary.get("笔记ID"))
    row["platform"] = "douyin" if platform == "dy" else "xiaohongshu"
    row["source_url"] = canonical_link(platform, note_id, summary.get("笔记链接"))
    row["note_id"] = note_id
    if platform == "dy":
        row["aweme_id"] = note_id
    for field in CORE_FIELDS:
        row[f"repair.{field}"] = interaction_total(summary) if field == "互动量" else clean_text(summary.get(field))
    comment = comment_index.get(note_id)
    if comment:
        row["repair.comment_rows"] = str(comment.get("commentRows") or "")
        row["repair.comment_file"] = " | ".join(comment.get("files") or [])
    row["repair.updated_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    return row


def make_backup(path: Path) -> str:
    backup_dir = PIPELINE_DIR / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{time.strftime('%Y%m%d_%H%M%S')}_{path.name}.bak"
    if path.exists():
        shutil.copy2(path, backup_path)
        return str(backup_path)
    return ""


def merge_fields(fields: list[str], rows: list[dict[str, Any]], new_fields: list[str]) -> list[str]:
    out = list(fields)
    for field in new_fields:
        if field not in out:
            out.append(field)
    for row in rows:
        for field in out:
            row.setdefault(field, "")
    return out


def has_missing_fields(row: dict[str, Any], fields: list[str]) -> list[str]:
    return [field for field in fields if is_blank(row.get(field))]


def row_text_blob(row: dict[str, Any], limit: int = 60000) -> str:
    chunks: list[str] = []
    for value in row.values():
        text = decode_urlish(str(value or ""))
        if text:
            chunks.append(text)
        if sum(len(item) for item in chunks) > limit:
            break
    return "\n".join(chunks)[:limit]


def find_tokenized_xhs_url(note_id: str, *row_groups: list[dict[str, Any]]) -> str:
    if not note_id:
        return ""
    pattern = re.compile(
        rf"https?://(?:www\.)?xiaohongshu\.com/(?:discovery/item|explore|search_result)/{re.escape(note_id)}[^\s\"'<>]*xsec_token=[^\s\"'<>]+",
        re.I,
    )
    for rows in row_groups:
        for row in rows:
            blob = row_text_blob(row)
            for match in pattern.finditer(blob):
                url = match.group(0).rstrip("。；;，,）)】]}")
                if "xsec_token=" in url:
                    return url.replace("&amp;", "&").replace("\\u0026", "&").replace("\\/", "/")
    return ""


def best_refresh_url(
    platform: str,
    note_id: str,
    row: dict[str, Any],
    origin_rows: list[dict[str, Any]],
    table_rows: list[dict[str, Any]],
) -> str:
    current = extract_platform_url(platform, row.get("笔记链接")) or extract_platform_url(platform, row_text_blob(row, 12000))
    if platform == "xhs":
        tokenized = find_tokenized_xhs_url(note_id, [row], origin_rows, table_rows)
        if tokenized:
            return tokenized
    if current:
        return current
    return canonical_link(platform, note_id, "")


def refresh_label(platform: str, note_id: str) -> str:
    safe_id = re.sub(r"[^\w.-]+", "_", note_id or "unknown").strip("_")
    return f"{platform}_{safe_id}_{time.strftime('%Y%m%d_%H%M%S')}_{int(time.time() * 1000) % 1000:03d}"


def run_note_refresh(
    platform: str,
    url: str,
    note_id: str,
    headed: bool,
    no_media_enrich: bool,
    timeout: float,
) -> tuple[Path, Path, str]:
    REFRESH_TMP_DIR.mkdir(parents=True, exist_ok=True)
    label = refresh_label(platform, note_id)
    origin_output = REFRESH_TMP_DIR / f"{label}_origin_data.csv"
    summary_output = REFRESH_TMP_DIR / f"{label}_10_fields.csv"
    script = NOTE_SCRIPTS[platform]
    command = [
        SCRIPT_PYTHON,
        str(script),
        url,
        "--output",
        str(origin_output),
        "--summary-output",
        str(summary_output),
    ]
    if platform == "xhs":
        command.append("--use-default-profile")
    if headed:
        command.append("--headed")
    if no_media_enrich:
        command.append("--no-media-enrich")
    print(f"[{PLATFORMS[platform]['name']}] 回源重爬：note_id={note_id} url={url}", flush=True)
    result = subprocess.run(
        command,
        cwd=str(PROJECT_ROOT),
        text=True,
        capture_output=True,
        check=False,
        timeout=timeout,
    )
    output_text = "\n".join(part for part in (result.stdout.strip(), result.stderr.strip()) if part)
    if result.returncode != 0:
        raise RuntimeError(output_text or f"crawler exited with {result.returncode}")
    if not origin_output.exists() or origin_output.stat().st_size == 0:
        raise RuntimeError("回源脚本未生成 origin_data 临时文件")
    return origin_output, summary_output, output_text


def merge_source_csv_into_rows(
    target_fields: list[str],
    target_rows: list[dict[str, Any]],
    source_path: Path,
) -> tuple[list[str], int]:
    source_fields, source_rows = read_csv(source_path)
    if not source_fields or not source_rows:
        return target_fields, 0
    target_fields = merge_fields(target_fields, target_rows, source_fields)
    for row in source_rows:
        merged = {field: "" for field in target_fields}
        for field in source_fields:
            merged[field] = row.get(field, "")
        target_rows.append(merged)
    return target_fields, len(source_rows)


def refresh_missing_rows(
    platform: str,
    table_rows: list[dict[str, Any]],
    origin_fields: list[str],
    origin_rows: list[dict[str, Any]],
    origin_index: dict[str, dict[str, Any]],
    missing_fields: list[str],
    max_refresh: int,
    headed: bool,
    no_media_enrich: bool,
    request_interval: float,
    jitter: float,
    retry: int,
    timeout: float,
    dry_run: bool = False,
) -> tuple[list[str], dict[str, Any]]:
    config = PLATFORMS[platform]
    candidates: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for row_index, row in enumerate(table_rows, start=2):
        note_id = clean_text(row.get("笔记ID")) or extract_note_id(platform, row.get("笔记链接")) or extract_note_id(platform, row_text_blob(row, 16000))
        if not note_id or note_id in seen_ids:
            continue
        missing = has_missing_fields(row, missing_fields)
        origin_missing = has_missing_fields(origin_index.get(note_id, {}), missing_fields)
        if missing or origin_missing:
            candidates.append({
                "row": row_index,
                "note_id": note_id,
                "missing": sorted(set(missing + origin_missing)),
                "row_data": row,
            })
            seen_ids.add(note_id)

    limit = len(candidates) if max_refresh <= 0 else min(max_refresh, len(candidates))
    report: dict[str, Any] = {
        "refreshCandidates": len(candidates),
        "refreshLimit": limit,
        "refreshAttempted": 0,
        "refreshSucceeded": 0,
        "refreshFailed": 0,
        "refreshErrors": [],
        "refreshExamples": [],
        "refreshCandidateSample": [
            {key: value for key, value in item.items() if key != "row_data"}
            for item in candidates[:20]
        ],
    }
    if dry_run:
        report["refreshDryRun"] = True
        return origin_fields, report
    if limit <= 0:
        return origin_fields, report

    for offset, candidate in enumerate(candidates[:limit], start=1):
        note_id = candidate["note_id"]
        url = best_refresh_url(platform, note_id, candidate["row_data"], origin_rows, table_rows)
        if not url:
            report["refreshFailed"] += 1
            report["refreshErrors"].append(f"{note_id}: 未找到可回源链接")
            continue
        if platform == "xhs" and "xsec_token=" not in url:
            print(
                f"[{config['name']}] note_id={note_id} 未找到历史 xsec_token 链接，将尝试页面解析；"
                "若账号处于 300013 风控状态建议稍后重试或粘贴 webshare 链接。",
                flush=True,
            )

        last_error = ""
        for attempt in range(1, max(1, retry) + 1):
            report["refreshAttempted"] += 1
            try:
                origin_output, summary_output, log_text = run_note_refresh(
                    platform,
                    url,
                    note_id,
                    headed=headed,
                    no_media_enrich=no_media_enrich,
                    timeout=timeout,
                )
                origin_fields, appended = merge_source_csv_into_rows(origin_fields, origin_rows, origin_output)
                _, summaries = read_csv(summary_output)
                summary = summaries[0] if summaries else summarize_origin_row(platform, origin_rows[-1], config["summary"])
                refreshed_id = clean_text(summary.get("笔记ID")) or note_id
                origin_index[refreshed_id] = merge_summary(origin_index.get(refreshed_id, {}), summary)
                for table_row in table_rows:
                    row_id = clean_text(table_row.get("笔记ID")) or extract_note_id(platform, table_row.get("笔记链接"))
                    if row_id != refreshed_id:
                        continue
                    for field in CORE_FIELDS:
                        value = interaction_total(summary) if field == "互动量" else clean_text(summary.get(field))
                        if value and (is_blank(table_row.get(field)) or (field == "笔记内容" and len(value) > len(clean_text(table_row.get(field))))):
                            table_row[field] = value
                    table_row["渠道类型"] = table_row.get("渠道类型") or config["channel"]
                report["refreshSucceeded"] += 1
                if len(report["refreshExamples"]) < 10:
                    report["refreshExamples"].append({
                        "note_id": refreshed_id,
                        "row": candidate["row"],
                        "url": url,
                        "appendedOriginRows": appended,
                        "title": clean_text(summary.get("笔记标题"))[:120],
                        "contentPreview": clean_text(summary.get("笔记内容"))[:180],
                        "tempOrigin": str(origin_output),
                        "tempSummary": str(summary_output),
                        "log": clean_text(log_text)[:300],
                    })
                break
            except Exception as exc:
                last_error = str(exc)
                if attempt < max(1, retry):
                    sleep_seconds = max(3.0, request_interval / 2) + random.uniform(0, max(0.0, jitter))
                    print(f"[{config['name']}] {note_id} 第 {attempt} 次回源失败，{sleep_seconds:.1f}s 后重试：{last_error}", flush=True)
                    time.sleep(sleep_seconds)
        else:
            report["refreshFailed"] += 1
            if len(report["refreshErrors"]) < 30:
                report["refreshErrors"].append(f"{note_id}: {last_error}")

        if offset < limit:
            sleep_seconds = max(0.0, request_interval) + random.uniform(0, max(0.0, jitter))
            print(f"[{config['name']}] 反爬保护等待 {sleep_seconds:.1f}s 后继续下一条。", flush=True)
            time.sleep(sleep_seconds)

    return origin_fields, report


def repair_platform(
    platform: str,
    dry_run: bool,
    recompute_interactions: bool,
    refresh_missing: bool = False,
    missing_fields: list[str] | None = None,
    max_refresh: int = 20,
    headed: bool = False,
    no_media_enrich: bool = False,
    request_interval: float = 18.0,
    jitter: float = 7.0,
    retry: int = 1,
    refresh_timeout: float = 180.0,
) -> dict[str, Any]:
    config = PLATFORMS[platform]
    table_path = Path(config["table"])
    origin_path = Path(config["origin"])
    table_fields, table_rows = read_csv(table_path)
    origin_fields, origin_rows = read_csv(origin_path)
    table_fields, table_schema_changed = ensure_fields(table_fields, table_rows, CORE_FIELDS)
    origin_required = ["platform", "source_url", "note_id"] + (["aweme_id"] if platform == "dy" else []) + ORIGIN_REPAIR_FIELDS
    origin_fields, origin_schema_changed = ensure_fields(origin_fields, origin_rows, origin_required)

    comment_index = build_comment_index(platform)

    origin_index: dict[str, dict[str, Any]] = {}
    origin_changed_rows = 0
    origin_field_updates: dict[str, int] = defaultdict(int)
    origin_examples: list[dict[str, Any]] = []

    for row_index, origin in enumerate(origin_rows, start=2):
        summary = summarize_origin_row(platform, origin, config["summary"])
        note_id = clean_text(summary.get("笔记ID"))
        if note_id:
            old = origin_index.get(note_id)
            if old is None or candidate_score(summary) > candidate_score(old):
                origin_index[note_id] = summary
        changed, changed_fields = update_origin_row(platform, origin, summary, comment_index)
        if changed:
            origin_changed_rows += 1
            for field in changed_fields:
                origin_field_updates[field] += 1
            if len(origin_examples) < 10:
                origin_examples.append({
                    "row": row_index,
                    "note_id": note_id,
                    "updated_fields": changed_fields[:12],
                    "title": clean_text(summary.get("笔记标题"))[:100],
                })

    table_changed_rows = 0
    table_field_updates: dict[str, int] = defaultdict(int)
    table_examples: list[dict[str, Any]] = []
    unresolved: list[dict[str, Any]] = []

    for row_index, row in enumerate(table_rows, start=2):
        note_id = clean_text(row.get("笔记ID")) or extract_note_id(platform, row.get("笔记链接"))
        link = canonical_link(platform, note_id, row.get("笔记链接"))
        changed_fields: list[str] = []
        if note_id and clean_text(row.get("笔记ID")) != note_id:
            row["笔记ID"] = note_id
            changed_fields.append("笔记ID")
            table_field_updates["笔记ID"] += 1
        if link and clean_text(row.get("笔记链接")) != link and not link.startswith(clean_text(row.get("笔记链接"))):
            row["笔记链接"] = link
            changed_fields.append("笔记链接")
            table_field_updates["笔记链接"] += 1

        candidate = dict(origin_index.get(note_id) or {})
        candidate = merge_summary(candidate, fallback_summary_from_row(platform, row))
        comment = comment_index.get(note_id)
        if comment and number_value(candidate.get("评论量")) == 0:
            candidate["评论量"] = str(comment.get("commentRows") or "")
            candidate["互动量"] = interaction_total(candidate)

        for field in CORE_FIELDS:
            if field == "互动量":
                if recompute_interactions or is_blank(row.get(field)):
                    merged_for_total = dict(candidate)
                    for key, value_item in row.items():
                        if not is_blank(value_item):
                            merged_for_total[key] = value_item
                    value = interaction_total(merged_for_total)
                    if clean_text(row.get(field)) != value:
                        row[field] = value
                        changed_fields.append(field)
                        table_field_updates[field] += 1
                continue
            if field == "渠道类型" and is_blank(row.get(field)):
                row[field] = config["channel"]
                changed_fields.append(field)
                table_field_updates[field] += 1
                continue
            if is_blank(row.get(field)) and not is_blank(candidate.get(field)):
                row[field] = clean_text(candidate[field])
                changed_fields.append(field)
                table_field_updates[field] += 1

        if changed_fields:
            table_changed_rows += 1
            if len(table_examples) < 12:
                table_examples.append({
                    "row": row_index,
                    "note_id": note_id,
                    "updated_fields": changed_fields,
                    "title": clean_text(row.get("笔记标题"))[:120],
                })

        missing = [field for field in CORE_FIELDS if field != "互动量" and is_blank(row.get(field))]
        if missing and len(unresolved) < 30:
            unresolved.append({
                "row": row_index,
                "note_id": note_id,
                "missing": missing,
                "link": row.get("笔记链接", ""),
            })

    existing_origin_ids = set(origin_index)
    origin_appended_rows = 0
    for row in table_rows:
        summary = merge_summary({}, fallback_summary_from_row(platform, row))
        note_id = clean_text(summary.get("笔记ID"))
        if not note_id or note_id in existing_origin_ids:
            continue
        origin_rows.append(build_minimal_origin_row(platform, origin_fields, summary, comment_index))
        origin_index[note_id] = summary
        existing_origin_ids.add(note_id)
        origin_appended_rows += 1

    refresh_report: dict[str, Any] = {
        "refreshCandidates": 0,
        "refreshLimit": 0,
        "refreshAttempted": 0,
        "refreshSucceeded": 0,
        "refreshFailed": 0,
    }
    refresh_origin_rows_before = len(origin_rows)
    if refresh_missing:
        origin_fields, refresh_report = refresh_missing_rows(
            platform=platform,
            table_rows=table_rows,
            origin_fields=origin_fields,
            origin_rows=origin_rows,
            origin_index=origin_index,
            missing_fields=missing_fields or DEFAULT_MISSING_FIELDS,
            max_refresh=max_refresh,
            headed=headed,
            no_media_enrich=no_media_enrich,
            request_interval=request_interval,
            jitter=jitter,
            retry=retry,
            timeout=refresh_timeout,
            dry_run=dry_run,
        )
        origin_appended_rows += max(0, len(origin_rows) - refresh_origin_rows_before)
        if refresh_report.get("refreshSucceeded"):
            # 回源成功后再跑一遍轻量本地补齐，确保新增 origin 字段写回 repair.*。
            for origin in origin_rows[refresh_origin_rows_before:]:
                summary = summarize_origin_row(platform, origin, config["summary"])
                note_id = clean_text(summary.get("笔记ID"))
                if note_id:
                    origin_index[note_id] = merge_summary(origin_index.get(note_id, {}), summary)
                changed, changed_fields = update_origin_row(platform, origin, summary, comment_index)
                if changed:
                    origin_changed_rows += 1
                    for field in changed_fields:
                        origin_field_updates[field] += 1
            for row in table_rows:
                note_id = clean_text(row.get("笔记ID")) or extract_note_id(platform, row.get("笔记链接"))
                candidate = origin_index.get(note_id) or {}
                changed_fields = []
                for field in CORE_FIELDS:
                    value = interaction_total(candidate) if field == "互动量" else clean_text(candidate.get(field))
                    if value and (is_blank(row.get(field)) or (field == "笔记内容" and len(value) > len(clean_text(row.get(field))))):
                        row[field] = value
                        changed_fields.append(field)
                        table_field_updates[field] += 1
                if changed_fields:
                    table_changed_rows += 1
                    if len(table_examples) < 12:
                        table_examples.append({
                            "row": "refresh",
                            "note_id": note_id,
                            "updated_fields": changed_fields,
                            "title": clean_text(row.get("笔记标题"))[:120],
                        })

    table_backup = ""
    origin_backup = ""
    if not dry_run:
        if table_rows and (table_changed_rows or table_schema_changed):
            table_backup = make_backup(table_path)
            write_csv(table_path, table_fields, table_rows)
        if origin_rows and (origin_changed_rows or origin_appended_rows or origin_schema_changed or refresh_report.get("refreshSucceeded")):
            origin_backup = make_backup(origin_path)
            write_csv(origin_path, origin_fields, origin_rows)

    final_unresolved: list[dict[str, Any]] = []
    for row_index, row in enumerate(table_rows, start=2):
        note_id = clean_text(row.get("笔记ID")) or extract_note_id(platform, row.get("笔记链接"))
        missing = [field for field in CORE_FIELDS if field != "互动量" and is_blank(row.get(field))]
        if missing and len(final_unresolved) < 30:
            final_unresolved.append({
                "row": row_index,
                "note_id": note_id,
                "missing": missing,
                "link": row.get("笔记链接", ""),
            })

    return {
        "platform": platform,
        "platformName": config["name"],
        "table": str(table_path),
        "origin": str(origin_path),
        "rows": len(table_rows),
        "originRows": len(origin_rows),
        "originIndexed": len(origin_index),
        "commentIndexed": len(comment_index),
        "tableChangedRows": table_changed_rows,
        "originChangedRows": origin_changed_rows,
        "originAppendedRows": origin_appended_rows,
        "refresh": refresh_report,
        "tableFieldUpdates": dict(table_field_updates),
        "originFieldUpdates": dict(origin_field_updates),
        "unresolvedSampleBeforeRefresh": unresolved,
        "unresolvedSample": final_unresolved,
        "tableExamples": table_examples,
        "originExamples": origin_examples,
        "tableBackup": table_backup,
        "originBackup": origin_backup,
        "dryRun": dry_run,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="检查并补全小红书/抖音监控总表与 origin_data 核心字段。")
    parser.add_argument("--platform", choices=["all", "xhs", "dy"], default="all")
    parser.add_argument("--dry-run", action="store_true", help="只输出报告，不写回 CSV。")
    parser.add_argument("--recompute-interactions", action="store_true", help="重新计算互动量。")
    parser.add_argument(
        "--refresh-missing",
        action="store_true",
        help="对本地仍缺核心字段的帖子逐条回源重爬，成功后写回 origin_data 和监控总表。",
    )
    parser.add_argument(
        "--missing-fields",
        default=",".join(DEFAULT_MISSING_FIELDS),
        help="判定需要回源重爬的核心字段，逗号分隔。",
    )
    parser.add_argument("--max-refresh", type=int, default=20, help="每个平台本次最多回源条数；0 表示不限制。")
    parser.add_argument("--headed", action="store_true", help="回源重爬时显示浏览器窗口，便于复用登录态和降低风控。")
    parser.add_argument("--no-media-enrich", action="store_true", help="回源重爬时跳过图片 OCR 和视频语音转文字。")
    parser.add_argument("--request-interval", type=float, default=18.0, help="回源重爬每条之间的基础等待秒数。")
    parser.add_argument("--jitter", type=float, default=7.0, help="回源重爬每条之间的随机附加等待秒数。")
    parser.add_argument("--retry", type=int, default=1, help="单条回源失败后的重试次数。")
    parser.add_argument("--refresh-timeout", type=float, default=180.0, help="单条回源脚本最长运行秒数。")
    parser.add_argument("--report", help="JSON 报告输出路径。")
    args = parser.parse_args()

    platforms = ["xhs", "dy"] if args.platform == "all" else [args.platform]
    missing_fields = [item.strip() for item in args.missing_fields.split(",") if item.strip()]
    results = [
        repair_platform(
            item,
            dry_run=args.dry_run,
            recompute_interactions=args.recompute_interactions,
            refresh_missing=args.refresh_missing,
            missing_fields=missing_fields,
            max_refresh=args.max_refresh,
            headed=args.headed,
            no_media_enrich=args.no_media_enrich,
            request_interval=args.request_interval,
            jitter=args.jitter,
            retry=args.retry,
            refresh_timeout=args.refresh_timeout,
        )
        for item in platforms
    ]
    payload = {
        "generatedAt": time.strftime("%Y-%m-%d %H:%M:%S"),
        "project": str(PROJECT_ROOT),
        "results": results,
    }
    report_path = Path(args.report) if args.report else PIPELINE_DIR / f"monitoring_repair_report_{time.strftime('%Y%m%d_%H%M%S')}.json"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({**payload, "report": str(report_path)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
