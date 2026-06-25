#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Local data server for the public opinion visualization dashboard.

Run:
  python3 Visualization/server.py --port 8765
  python3 Visualization/server.py --check
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import mimetypes
import os
import re
import subprocess
import sys
import time
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse
from xml.etree import ElementTree as ET


VIS_ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = VIS_ROOT.parent
PIPELINE_DIR = PROJECT_ROOT / "Pipeline"
COMMENT_DIR = PROJECT_ROOT / "Comment_Data"
HYPE_DIR = PROJECT_ROOT / "Hype_Something"
EXTERNAL_DATA_DIR = PROJECT_ROOT / "External_Data"

DATA_TABLES = {
    "xhs": PIPELINE_DIR / "xhs_Data_Table_on_Channel_Public_Opinion_Monitoring_2026.csv",
    "dy": PIPELINE_DIR / "dy_Data_Table_on_Channel_Public_Opinion_Monitoring_2026.csv",
}
COMMENT_DIRS = {
    "xhs": COMMENT_DIR / "xhs",
    "dy": COMMENT_DIR / "dy",
}
HYPE_WORKBOOKS = {
    "xhs": HYPE_DIR / "2026_Didi_Xiaohongshu_Daily_Word-of-Mouth_Amplification.xlsx",
    "dy": HYPE_DIR / "2026_Didi_Douyin_Daily_Word-of-Mouth_Amplification.xlsx",
}

PLATFORM_LABELS = {"xhs": "小红书", "dy": "抖音"}
STOPWORDS = {
    "一个", "一下", "一些", "这个", "那个", "我们", "他们", "你们", "自己", "没有", "不是", "还是", "真的",
    "感觉", "就是", "因为", "所以", "如果", "但是", "然后", "已经", "可以", "不能", "不会", "小红书",
    "抖音", "滴滴", "出行", "打车", "司机", "平台", "用户", "视频", "评论", "内容", "话题", "分享",
}


def candidate_pythons() -> list[str]:
    raw = [
        os.environ.get("PIPELINE_PYTHON", ""),
        str(PROJECT_ROOT / ".venv" / "bin" / "python"),
        "/Library/Developer/CommandLineTools/usr/bin/python3",
        "/usr/bin/python3",
        sys.executable,
        "python3",
    ]
    out: list[str] = []
    seen: set[str] = set()
    for item in raw:
        if not item or item in seen:
            continue
        if "/" in item and not Path(item).exists():
            continue
        out.append(item)
        seen.add(item)
    return out


def can_import(python_bin: str, module: str) -> bool:
    try:
        result = subprocess.run(
            [python_bin, "-c", f"import {module}"],
            cwd=str(PROJECT_ROOT),
            text=True,
            capture_output=True,
            timeout=8,
            check=False,
        )
        return result.returncode == 0
    except Exception:
        return False


def can_import_all(python_bin: str, modules: list[str]) -> bool:
    return all(can_import(python_bin, module) for module in modules)


def has_pdf_parser(python_bin: str) -> bool:
    return can_import(python_bin, "pypdf") or can_import(python_bin, "pdfplumber") or can_import(python_bin, "fitz")


def reexec_with_openpyxl_if_available() -> None:
    try:
        import openpyxl  # noqa: F401
        current_has_openpyxl = True
    except Exception:
        current_has_openpyxl = False
    try:
        import pypdf  # type: ignore  # noqa: F401
        current_has_pdf = True
    except Exception:
        try:
            import pdfplumber  # type: ignore  # noqa: F401
            current_has_pdf = True
        except Exception:
            try:
                import fitz  # type: ignore  # noqa: F401
                current_has_pdf = True
            except Exception:
                current_has_pdf = False
    if current_has_openpyxl and current_has_pdf:
        return
    candidates = candidate_pythons()
    for python_bin in candidates:
        try:
            same = Path(python_bin).resolve() == Path(sys.executable).resolve()
        except Exception:
            same = False
        if same:
            continue
        if can_import_all(python_bin, ["openpyxl"]) and has_pdf_parser(python_bin):
            os.execv(python_bin, [python_bin] + sys.argv)
    if current_has_openpyxl:
        return
    for python_bin in candidates:
        try:
            same = Path(python_bin).resolve() == Path(sys.executable).resolve()
        except Exception:
            same = False
        if same:
            continue
        if can_import_all(python_bin, ["openpyxl"]):
            os.execv(python_bin, [python_bin] + sys.argv)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def is_blank(value: Any) -> bool:
    return clean_text(value).lower() in {"", "none", "null", "nan", "undefined"}


def number_value(value: Any) -> int:
    text = clean_text(value).replace(",", "")
    if not text:
        return 0
    for suffix, factor in (("万", 10000), ("千", 1000), ("w", 10000), ("W", 10000), ("k", 1000), ("K", 1000)):
        if text.endswith(suffix):
            try:
                return int(float(text[: -len(suffix)]) * factor)
            except Exception:
                return 0
    try:
        return int(float(text))
    except Exception:
        return 0


def parse_time(value: Any) -> datetime | None:
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{10,13}", text):
        try:
            timestamp = int(text)
            if timestamp > 10**12:
                timestamp = timestamp // 1000
            return datetime.fromtimestamp(timestamp)
        except Exception:
            return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:19], fmt)
        except Exception:
            pass
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except Exception:
            return None
    return None


def month_key(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m") if dt else ""


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_posts() -> list[dict[str, Any]]:
    posts: list[dict[str, Any]] = []
    for platform, path in DATA_TABLES.items():
        for row in read_csv_rows(path):
            note_id = clean_text(row.get("笔记ID"))
            publish_dt = parse_time(row.get("发布时间"))
            likes = number_value(row.get("点赞量"))
            collects = number_value(row.get("收藏量"))
            comments = number_value(row.get("评论量"))
            shares = number_value(row.get("分享量"))
            interactions = number_value(row.get("互动量")) or (likes + collects + comments + shares)
            posts.append({
                "platform": platform,
                "platformName": PLATFORM_LABELS[platform],
                "noteId": note_id,
                "publishTime": clean_text(row.get("发布时间")),
                "publishDate": publish_dt.strftime("%Y-%m-%d") if publish_dt else "",
                "month": month_key(publish_dt),
                "title": clean_text(row.get("笔记标题")),
                "url": clean_text(row.get("笔记链接")),
                "content": clean_text(row.get("笔记内容")),
                "author": clean_text(row.get("博主昵称")),
                "likes": likes,
                "collects": collects,
                "comments": comments,
                "shares": shares,
                "interactions": interactions,
                "summary": clean_text(row.get("概括")),
                "contentType": clean_text(row.get("内容类型")) or "未填写",
                "sentiment": clean_text(row.get("正负向")) or "未填写",
                "businessLine": clean_text(row.get("业务线")) or "未填写",
                "scenario": clean_text(row.get("具体产品/场景")) or "未填写",
            })
    return posts


def openpyxl_module() -> Any | None:
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception:
        return None


def read_xlsx_rows(path: Path, max_rows: int = 0) -> list[dict[str, str]]:
    openpyxl = openpyxl_module()
    if not path.exists():
        return []
    if openpyxl is None:
        return read_xlsx_rows_zip(path, max_rows=max_rows)
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        iterator = ws.iter_rows(values_only=True)
        headers = [clean_text(cell) for cell in next(iterator, [])]
        rows: list[dict[str, str]] = []
        for index, values in enumerate(iterator, start=1):
            if max_rows and index > max_rows:
                break
            item = {headers[i]: clean_text(values[i]) if i < len(values) else "" for i in range(len(headers))}
            if any(item.values()):
                rows.append(item)
        wb.close()
        return rows
    except Exception:
        return read_xlsx_rows_zip(path, max_rows=max_rows)


def column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref or "")
    if not match:
        return 0
    value = 0
    for char in match.group(1):
        value = value * 26 + (ord(char) - ord("A") + 1)
    return max(value - 1, 0)


def read_shared_strings(zf: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    except Exception:
        return []
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    values: list[str] = []
    for item in root.findall(f".//{ns}si"):
        values.append("".join(node.text or "" for node in item.findall(f".//{ns}t")))
    return values


def xlsx_cell_value(cell: ET.Element, shared: list[str]) -> str:
    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    cell_type = cell.attrib.get("t", "")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(f".//{ns}t"))
    value_node = cell.find(f"{ns}v")
    raw = value_node.text if value_node is not None else ""
    if cell_type == "s":
        try:
            return shared[int(raw)]
        except Exception:
            return ""
    return raw or ""


def read_xlsx_rows_zip(path: Path, max_rows: int = 0) -> list[dict[str, str]]:
    try:
        with zipfile.ZipFile(path) as zf:
            sheet_names = sorted(name for name in zf.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name))
            if not sheet_names:
                return []
            shared = read_shared_strings(zf)
            root = ET.fromstring(zf.read(sheet_names[0]))
            ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
            raw_rows: list[list[str]] = []
            for row in root.findall(f".//{ns}sheetData/{ns}row"):
                values: list[str] = []
                for cell in row.findall(f"{ns}c"):
                    index = column_index(cell.attrib.get("r", ""))
                    while len(values) <= index:
                        values.append("")
                    values[index] = clean_text(xlsx_cell_value(cell, shared))
                if any(values):
                    raw_rows.append(values)
            if not raw_rows:
                return []
            headers = raw_rows[0]
            rows: list[dict[str, str]] = []
            for index, values in enumerate(raw_rows[1:], start=1):
                if max_rows and index > max_rows:
                    break
                item = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}
                if any(item.values()):
                    rows.append(item)
            return rows
    except Exception:
        return []


def cell_text(value: Any) -> str:
    text = clean_text(value)
    if text in {"[object Object]", "None", "none", "null", "NULL"}:
        return ""
    return text


def read_workbook_rows(path: Path, max_rows_per_sheet: int = 800) -> list[dict[str, Any]]:
    openpyxl = openpyxl_module()
    if openpyxl is None or not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            iterator = ws.iter_rows(values_only=True)
            headers = [cell_text(cell) for cell in next(iterator, [])]
            if not any(headers):
                continue
            for row_index, values in enumerate(iterator, start=2):
                if row_index > max_rows_per_sheet:
                    break
                item = {headers[i] or f"字段{i + 1}": cell_text(values[i]) if i < len(values) else "" for i in range(len(headers))}
                if not any(item.values()):
                    continue
                if all(value == "汇总" or value == "-" for value in item.values() if value):
                    continue
                item["_sheet"] = ws.title
                item["_row"] = row_index
                rows.append(item)
    except Exception:
        return []
    return rows


def load_comments() -> list[dict[str, Any]]:
    comments: list[dict[str, Any]] = []
    for platform, directory in COMMENT_DIRS.items():
        if not directory.exists():
            continue
        for path in sorted(directory.glob("*.xlsx")):
            note_id = path.stem
            for row in read_xlsx_rows(path):
                content = clean_text(row.get("评论内容"))
                if not content:
                    continue
                comments.append({
                    "platform": platform,
                    "platformName": PLATFORM_LABELS[platform],
                    "noteId": clean_text(row.get("笔记ID")) or note_id,
                    "commentId": clean_text(row.get("评论ID")),
                    "content": content,
                    "likes": number_value(row.get("点赞量")),
                    "time": clean_text(row.get("评论时间")),
                    "ip": clean_text(row.get("IP地址")),
                    "user": clean_text(row.get("用户名称")),
                    "sourceFile": str(path.relative_to(PROJECT_ROOT)),
                })
        for path in sorted(directory.glob("*.csv")):
            note_id = path.stem
            for row in read_csv_rows(path):
                content = clean_text(row.get("评论内容"))
                if not content:
                    continue
                comments.append({
                    "platform": platform,
                    "platformName": PLATFORM_LABELS[platform],
                    "noteId": clean_text(row.get("笔记ID")) or note_id,
                    "commentId": clean_text(row.get("评论ID")),
                    "content": content,
                    "likes": number_value(row.get("点赞量")),
                    "time": clean_text(row.get("评论时间")),
                    "ip": clean_text(row.get("IP地址")),
                    "user": clean_text(row.get("用户名称")),
                    "sourceFile": str(path.relative_to(PROJECT_ROOT)),
                })
    return comments


def load_hype_summary() -> dict[str, Any]:
    openpyxl = openpyxl_module()
    result: dict[str, Any] = {}
    for platform, path in HYPE_WORKBOOKS.items():
        payload = {"path": str(path), "exists": path.exists(), "sheets": [], "rows": 0}
        if openpyxl is not None and path.exists():
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                payload["sheets"] = wb.sheetnames
                total = 0
                for ws in wb.worksheets:
                    total += max(ws.max_row - 1, 0)
                payload["rows"] = total
            except Exception as exc:
                payload["error"] = str(exc)
        result[platform] = payload
    return result


def pick_first(row: dict[str, Any], names: list[str]) -> str:
    for name in names:
        value = cell_text(row.get(name))
        if value:
            return value
    return ""


def money_from_record(text: str) -> int:
    total = 0.0
    for match in re.findall(r"投放\s*([0-9]+(?:\.[0-9]+)?)\s*元|([0-9]+(?:\.[0-9]+)?)\s*元", text or ""):
        value = next((part for part in match if part), "")
        if value:
            try:
                total += float(value)
            except Exception:
                pass
    return int(total)


def load_hype_details(selected_month: str) -> dict[str, Any]:
    month_number = ""
    try:
        month_number = str(int(selected_month.split("-")[1]))
    except Exception:
        pass
    all_rows: list[dict[str, Any]] = []
    totals = {"posts": 0, "spend": 0, "before": 0, "after": 0, "brought": 0}
    for platform, path in HYPE_WORKBOOKS.items():
        for row in read_workbook_rows(path):
            sheet = cell_text(row.get("_sheet"))
            if month_number and month_number not in sheet and f"{month_number}月" not in sheet:
                continue
            title = pick_first(row, ["标题", "笔记标题"])
            link = pick_first(row, ["链接", "笔记链接"])
            if not title and not link:
                continue
            before = number_value(pick_first(row, ["投前互动量", "投前笔记互动量"]))
            after = number_value(pick_first(row, ["投后笔记总互动量", "投后互动量", "总互动量", "投放后总互动量"]))
            brought = number_value(pick_first(row, ["投放带来的互动量", "投放互动量"])) or max(after - before, 0)
            record = pick_first(row, ["投放记录\nxx日投放xx元", "投放记录", "投放记录 xx日投放xx元"])
            spend = number_value(pick_first(row, ["总投放金额", "投放金额"])) or money_from_record(record)
            item = {
                "platform": PLATFORM_LABELS[platform],
                "sheet": sheet,
                "date": pick_first(row, ["发布日期", "发布时间"]),
                "status": pick_first(row, ["状态"]),
                "author": pick_first(row, ["作者昵称", "达人名称", "博主昵称"]),
                "title": title,
                "url": link,
                "noteId": pick_first(row, ["笔记ID", "视频ID"]),
                "before": before,
                "after": after,
                "brought": brought,
                "spend": spend,
                "cpe": round(spend / brought, 2) if spend and brought else "",
            }
            totals["posts"] += 1
            totals["spend"] += spend
            totals["before"] += before
            totals["after"] += after
            totals["brought"] += brought
            all_rows.append(item)
    all_rows.sort(key=lambda row: int(row.get("brought") or 0), reverse=True)
    return {
        "totals": {
            **totals,
            "cpe": round(totals["spend"] / totals["brought"], 2) if totals["spend"] and totals["brought"] else "",
        },
        "topRows": all_rows[:12],
    }


def extract_pdf_text(path: Path) -> tuple[str, str]:
    try:
        import pypdf  # type: ignore
        reader = pypdf.PdfReader(str(path))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        return text, "pypdf"
    except Exception:
        pass
    try:
        import pdfplumber  # type: ignore
        with pdfplumber.open(str(path)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        return text, "pdfplumber"
    except Exception:
        pass
    try:
        import fitz  # type: ignore
        doc = fitz.open(str(path))
        text = "\n".join(page.get_text() for page in doc)
        return text, "pymupdf"
    except Exception:
        pass
    return "", "missing_pdf_text_dependency"


def split_pdf_sections(text: str, fallback_title: str) -> list[dict[str, Any]]:
    lines = [clean_text(line) for line in text.splitlines()]
    lines = [line for line in lines if line]
    sections: list[dict[str, Any]] = []
    current = {"title": fallback_title, "body": []}
    heading_re = re.compile(r"^((第[一二三四五六七八九十]+[章节部分])|([一二三四五六七八九十]+[、.．])|([0-9]{1,2}[、.．])|([A-Z][、.．]))")
    for line in lines:
        is_heading = bool(heading_re.match(line)) or (len(line) <= 28 and not line.endswith(("。", "，", ",", "；", ";")) and len(sections) < 20)
        if is_heading and current["body"]:
            sections.append({"title": current["title"], "body": " ".join(current["body"])[:520]})
            current = {"title": line[:80], "body": []}
        elif is_heading and not current["body"]:
            current["title"] = line[:80]
        else:
            current["body"].append(line)
    if current["body"]:
        sections.append({"title": current["title"], "body": " ".join(current["body"])[:520]})
    if not sections and lines:
        sections.append({"title": fallback_title, "body": " ".join(lines[:20])[:520]})
    return sections[:12]


def load_external_reports() -> dict[str, Any]:
    pdfs = sorted(EXTERNAL_DATA_DIR.glob("**/*.pdf")) if EXTERNAL_DATA_DIR.exists() else []
    reports: list[dict[str, Any]] = []
    errors: list[str] = []
    aggregate_words: Counter[str] = Counter()
    for path in pdfs:
        text, engine = extract_pdf_text(path)
        if not text:
            errors.append(f"{path.relative_to(PROJECT_ROOT)}: 缺少 PDF 文本解析依赖，请安装 pypdf 或 pdfplumber")
            reports.append({
                "file": str(path.relative_to(PROJECT_ROOT)),
                "title": path.stem,
                "engine": engine,
                "sections": [],
            })
            continue
        sections = split_pdf_sections(text, path.stem)
        for section in sections:
            aggregate_words.update(content_tokens(section.get("title", "")))
            aggregate_words.update(content_tokens(section.get("body", "")))
        reports.append({
            "file": str(path.relative_to(PROJECT_ROOT)),
            "title": path.stem,
            "engine": engine,
            "sections": sections,
        })
    return {
        "files": len(pdfs),
        "reports": reports,
        "errors": errors,
        "keywords": [{"name": key, "value": value} for key, value in aggregate_words.most_common(12)],
    }


def filter_by_request(posts: list[dict[str, Any]], comments: list[dict[str, Any]], query: dict[str, list[str]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, str]:
    platform = query.get("platform", ["all"])[0]
    selected_month = query.get("month", [""])[0]
    compare_month = query.get("compare", [""])[0]
    if not selected_month:
        months = sorted({post["month"] for post in posts if post.get("month")})
        selected_month = months[-1] if months else datetime.now().strftime("%Y-%m")
    if not compare_month:
        try:
            year, month = [int(part) for part in selected_month.split("-")]
            month -= 1
            if month == 0:
                year -= 1
                month = 12
            compare_month = f"{year:04d}-{month:02d}"
        except Exception:
            compare_month = ""
    if platform in {"xhs", "dy"}:
        posts = [post for post in posts if post["platform"] == platform]
        comments = [comment for comment in comments if comment["platform"] == platform]
    month_posts = [post for post in posts if post.get("month") == selected_month] or posts
    note_ids = {post["noteId"] for post in month_posts if post.get("noteId")}
    month_comments = [comment for comment in comments if comment.get("noteId") in note_ids] if note_ids else comments
    return month_posts, month_comments, selected_month, compare_month


def attach_exported_comment_counts(posts: list[dict[str, Any]], comments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counter: Counter[str] = Counter()
    for comment in comments:
        note_id = clean_text(comment.get("noteId"))
        if note_id:
            counter[note_id] += 1
    enriched: list[dict[str, Any]] = []
    for post in posts:
        item = dict(post)
        exported = counter.get(clean_text(post.get("noteId")), 0)
        item["exportedComments"] = exported
        item["displayComments"] = exported or int(post.get("comments") or 0)
        enriched.append(item)
    return enriched


def daily_series(posts: list[dict[str, Any]], metric: str) -> list[dict[str, Any]]:
    bucket: dict[str, int] = defaultdict(int)
    for post in posts:
        day = post.get("publishDate") or "未识别"
        bucket[day] += int(post.get(metric) or 0)
    days = sorted(bucket)
    return [{"date": day, "value": bucket[day]} for day in days[-31:]]


def top_counter(items: list[str], limit: int = 10) -> list[dict[str, Any]]:
    counter = Counter(clean_text(item) for item in items if not is_blank(item))
    return [{"name": key, "value": value} for key, value in counter.most_common(limit)]


def content_tokens(text: str) -> list[str]:
    text = clean_text(text)
    tokens: list[str] = []
    for match in re.findall(r"[\u4e00-\u9fff]{2,8}", text):
        if match in STOPWORDS:
            continue
        if len(match) > 6:
            tokens.extend([match[i : i + 4] for i in range(0, len(match), 4)])
        else:
            tokens.append(match)
    return [token for token in tokens if token not in STOPWORDS and len(token) >= 2]


def word_cloud(posts: list[dict[str, Any]], comments: list[dict[str, Any]], limit: int = 40) -> list[dict[str, Any]]:
    counter: Counter[str] = Counter()
    for post in posts:
        for field in ("scenario", "businessLine", "contentType"):
            value = clean_text(post.get(field))
            if value and value != "未填写":
                counter[value] += 8
        counter.update(content_tokens(post.get("title", "")))
        counter.update(content_tokens(post.get("content", "")))
    for comment in comments:
        counter.update(content_tokens(comment.get("content", "")))
    return [{"text": key, "value": value} for key, value in counter.most_common(limit)]


def platform_table(posts: list[dict[str, Any]], comments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for platform in ("xhs", "dy"):
        platform_posts = [post for post in posts if post["platform"] == platform]
        platform_comments = [comment for comment in comments if comment["platform"] == platform]
        rows.append({
            "platform": PLATFORM_LABELS[platform],
            "posts": len(platform_posts),
            "comments": len(platform_comments) or sum(post.get("comments", 0) for post in platform_posts),
            "interactions": sum(post.get("interactions", 0) for post in platform_posts),
            "collects": sum(post.get("collects", 0) for post in platform_posts),
            "shares": sum(post.get("shares", 0) for post in platform_posts),
        })
    return rows


def sentiment_payload(posts: list[dict[str, Any]]) -> dict[str, Any]:
    counter = Counter(post.get("sentiment") or "未填写" for post in posts)
    total = sum(counter.values()) or 1
    positive = counter.get("正向", 0)
    negative = counter.get("负向", 0)
    return {
        "positiveRate": round(positive / total * 100, 1),
        "negativeRate": round(negative / total * 100, 1),
        "items": [{"name": key, "value": value} for key, value in counter.most_common()],
        "positiveTopics": top_counter([post.get("scenario", "") for post in posts if post.get("sentiment") == "正向"], 10),
        "negativeTopics": top_counter([post.get("scenario", "") for post in posts if post.get("sentiment") == "负向"], 10),
    }


def voices(posts: list[dict[str, Any]], limit: int = 8) -> list[dict[str, Any]]:
    candidates = sorted(posts, key=lambda item: int(item.get("interactions") or 0), reverse=True)
    out: list[dict[str, Any]] = []
    for post in candidates:
        text = clean_text(post.get("summary")) or clean_text(post.get("content")) or clean_text(post.get("title"))
        if not text:
            continue
        out.append({
            "author": post.get("author") or "匿名用户",
            "platform": post.get("platformName"),
            "date": post.get("publishDate"),
            "title": post.get("title"),
            "text": text[:96],
            "likes": post.get("likes"),
            "sentiment": post.get("sentiment"),
            "url": post.get("url"),
        })
        if len(out) >= limit:
            break
    return out


def event_rows(posts: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    ranked = sorted(posts, key=lambda item: int(item.get("interactions") or 0), reverse=True)
    rows: list[dict[str, Any]] = []
    for post in ranked[:limit]:
        rows.append({
            "event": post.get("title") or post.get("scenario") or "未命名事件",
            "type": post.get("contentType"),
            "discussions": post.get("displayComments") or post.get("comments"),
            "heat": min(5, max(1, math.ceil((int(post.get("interactions") or 0) + 1) / 50))),
            "status": "机会" if post.get("sentiment") == "正向" else ("风险" if post.get("sentiment") == "负向" else "观察"),
        })
    return rows


def build_dashboard_data(query: dict[str, list[str]] | None = None) -> dict[str, Any]:
    query = query or {}
    all_posts = load_posts()
    all_comments = load_comments()
    posts, comments, selected_month, compare_month = filter_by_request(all_posts, all_comments, query)
    posts = attach_exported_comment_counts(posts, comments)

    total_posts = len(posts)
    exported_comments = len(comments)
    total_comments = exported_comments or sum(post.get("comments", 0) for post in posts)
    total_interactions = sum(post.get("interactions", 0) for post in posts)
    total_collects = sum(post.get("collects", 0) for post in posts)
    total_shares = sum(post.get("shares", 0) for post in posts)

    content_type_items = top_counter([post.get("contentType", "") for post in posts], 8)
    scenario_items = top_counter([post.get("scenario", "") for post in posts], 10)
    business_items = top_counter([post.get("businessLine", "") for post in posts], 8)
    word_items = word_cloud(posts, comments)
    sentiment = sentiment_payload(posts)
    hype_details = load_hype_details(selected_month)
    external_reports = load_external_reports()

    positive_rate = sentiment["positiveRate"]
    negative_rate = sentiment["negativeRate"]
    summaries = [
        f"本月共监测 {total_posts} 条 UGC，累计互动 {total_interactions:,}，已导出评论 {exported_comments:,} 条。",
        f"正向占比 {positive_rate}% ，负向占比 {negative_rate}% ，高频场景集中在 {', '.join(item['name'] for item in scenario_items[:3]) or '暂无'}。",
        f"平台对比中，小红书与抖音合计覆盖 {len({post['noteId'] for post in posts if post.get('noteId')})} 个帖子 ID，评论区导出文件 {len(list((COMMENT_DIR / 'xhs').glob('*.xlsx'))) + len(list((COMMENT_DIR / 'dy').glob('*.xlsx')))} 个。",
        "建议优先跟进正向高互动内容、负向高评论内容，以及评论区重复出现的痛点词。",
    ]

    return {
        "generatedAt": now_text(),
        "month": selected_month,
        "compareMonth": compare_month,
        "sourcePaths": {
            "pipeline": str(PIPELINE_DIR),
            "comments": str(COMMENT_DIR),
            "hype": str(HYPE_DIR),
        },
        "totals": {
            "posts": total_posts,
            "comments": total_comments,
            "exportedComments": exported_comments,
            "interactions": total_interactions,
            "collects": total_collects,
            "shares": total_shares,
            "commentFiles": sum(1 for directory in COMMENT_DIRS.values() if directory.exists() for _ in directory.glob("*.xlsx")),
        },
        "summary": summaries,
        "trend": {
            "posts": daily_series(posts, "interactions"),
            "comments": daily_series(posts, "displayComments"),
        },
        "platforms": platform_table(posts, comments),
        "contentTypes": content_type_items,
        "scenarios": scenario_items,
        "businessLines": business_items,
        "sentiment": sentiment,
        "wordCloud": word_items,
        "voices": voices(posts),
        "events": event_rows(posts),
        "search": {
            "internal": {
                "destinationHeat": scenario_items[:7],
                "keywordHeat": word_items[:10],
                "intersection": business_items[:6],
            },
            "external": external_reports,
            "destinationHeat": scenario_items[:7],
            "keywordHeat": word_items[:10],
            "intersection": business_items[:6],
        },
        "industry": {
            "actions": event_rows(posts, 5),
            "categories": content_type_items,
        },
        "hot": {
            "events": event_rows(posts, 10),
            "types": content_type_items,
        },
        "monthlyChanges": {
            "cards": [
                {"title": "发布与讨论", "text": f"新增监测 {total_posts} 条，互动 {total_interactions:,}。"},
                {"title": "搜索趋势", "text": f"高频场景：{', '.join(item['name'] for item in scenario_items[:3]) or '暂无'}。"},
                {"title": "竞品行业事件", "text": f"内容类型覆盖 {len(content_type_items)} 类，需关注投诉与产品力变化。"},
                {"title": "热点事件", "text": f"高互动事件 {len(event_rows(posts, 10))} 个，风险类需单独复核。"},
                {"title": "下月建议", "text": "持续补齐评论区数据，优先沉淀正向口碑素材与负向问题闭环。"},
            ]
        },
        "hype": {
            "summary": load_hype_summary(),
            "details": hype_details,
        },
    }


class DashboardHandler(BaseHTTPRequestHandler):
    server_version = "PublicOpinionVisualization/1.0"

    def log_message(self, format: str, *args: Any) -> None:
        return

    def send_json(self, payload: Any, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, path: Path) -> None:
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        body = path.read_bytes()
        content_type = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        if path.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        elif path.suffix in {".html", ".css"}:
            content_type = f"text/{path.suffix[1:]}; charset=utf-8"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            self.send_json({"ok": True, "generatedAt": now_text(), "project": str(PROJECT_ROOT)})
            return
        if parsed.path == "/api/dashboard-data":
            try:
                self.send_json(build_dashboard_data(parse_qs(parsed.query)))
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return
        relative = unquote(parsed.path.lstrip("/")) or "index.html"
        if relative == "":
            relative = "index.html"
        target = (VIS_ROOT / relative).resolve()
        if not str(target).startswith(str(VIS_ROOT.resolve())):
            self.send_error(403)
            return
        if target.is_dir():
            target = target / "index.html"
        self.send_file(target)


def main() -> int:
    parser = argparse.ArgumentParser(description="Run the public opinion visualization dashboard.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--check", action="store_true", help="Print dashboard data summary and exit.")
    args = parser.parse_args()

    reexec_with_openpyxl_if_available()
    if args.check:
        data = build_dashboard_data({})
        print(json.dumps({
            "generatedAt": data["generatedAt"],
            "month": data["month"],
            "totals": data["totals"],
            "platforms": data["platforms"],
            "hype": data["hype"]["details"]["totals"],
            "externalReports": data["search"]["external"]["files"],
            "sourcePaths": data["sourcePaths"],
        }, ensure_ascii=False, indent=2))
        return 0

    httpd = ThreadingHTTPServer((args.host, args.port), DashboardHandler)
    print(f"Visualization dashboard: http://{args.host}:{args.port}/")
    print("Press Ctrl+C to stop.")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        httpd.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
