#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import importlib
import json
import math
import os
import re
import site
import subprocess
import sys
import time
import tempfile
import urllib.error
import urllib.request
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from pipeline_paths import XHS_DATA_TABLE_CSV, XHS_HYPE_WORKBOOK


ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = ROOT.parent
HYPE_ROOT = PROJECT_ROOT / "Hype_Something"
DATA_TABLE = XHS_DATA_TABLE_CSV
HYPE_HISTORY = HYPE_ROOT / "training_data_cleaned.csv"
HYPE_WORKBOOK = XHS_HYPE_WORKBOOK
HYPE_LLM_CONFIG = HYPE_ROOT / "llm_config.json"

MODEL_OPTIONS = [
    "All Named Models",
    "All Internal Models",
    "kimi-k2.5-external",
    "minimax-m2.5-external",
    "glm-5.1-external",
    "glm-5-external",
    "glm-5-internal",
    "glm-5.1-internal",
]
MODEL_GROUPS = {
    "All Named Models": ["kimi-k2.5-external", "minimax-m2.5-external", "glm-5.1-external", "glm-5-external"],
    "All Internal Models": ["glm-5-internal", "glm-5.1-internal"],
}

KEYWORDS = [
    "打车",
    "司机",
    "专车",
    "快车",
    "顺风车",
    "女性",
    "女司机",
    "安全",
    "通勤",
    "机场",
    "高铁",
    "夜晚",
    "服务",
    "体验",
    "真实",
    "避雷",
    "攻略",
    "省钱",
    "优惠",
    "暖心",
    "舒服",
    "惊喜",
    "故事",
    "日常",
    "测评",
    "宠物",
    "拼车",
    "巴士",
    "旅游",
    "马拉松",
    "香卡",
    "异味",
]
STOP_CHARS = set("的一是在不了和有就人都说而及与着或一个也很被到吧吗呢啊呀哦得比还更最太")

HEADER_ALIASES = {
    "publish_date": ["发布日期", "发布时间", "笔记日期", "发布日"],
    "end_date": ["投放截止", "截止日期", "结束日期", "投放结束"],
    "status": ["状态", "投放状态"],
    "author": ["作者昵称", "博主昵称", "昵称"],
    "link": ["链接", "笔记链接", "url"],
    "note_id": ["笔记ID", "笔记id", "noteid", "id"],
    "ad_records": ["投放记录xx日投放xx元", "投放记录", "投放明细", "投放过程"],
    "title": ["标题", "笔记标题", "题目"],
    "pre_interactions": ["投前互动量", "投前互动", "现有互动量", "当前互动量"],
    "post_total_interactions": ["投后笔记总互动量", "投后互动量", "总互动量"],
    "post_likes": ["投放后点赞量", "投放后点赞", "投后点赞"],
    "post_collects": ["投放后收藏量", "投放后收藏", "投后收藏"],
    "post_comments": ["投放后评论量", "投放后评论", "投后评论"],
    "post_increment_interactions": ["投放带来的互动量", "投放带来互动量"],
    "total_spend": ["总投放金额", "投放金额", "消耗金额", "总消耗"],
    "placement_cpe": ["投放CPE", "投放cpe"],
    "combined_cpe": ["综合cpe投入笔记外层互动数据", "综合cpe", "CPE"],
    "category1": ["一级分类", "内容类型", "分类1"],
    "category2": ["二级分类", "具体产品场景", "具体产品/场景", "分类2"],
    "remark": ["备注"],
    "viral_level": ["爆款定级"],
}

MONTH_NAMES = {
    1: ["1月", "一月"],
    2: ["2月", "二月"],
    3: ["3月", "三月"],
    4: ["4月", "四月"],
    5: ["5月", "五月"],
    6: ["6月", "六月"],
    7: ["7月", "七月"],
    8: ["8月", "八月"],
    9: ["9月", "九月"],
    10: ["10月", "十月"],
    11: ["11月", "十一月"],
    12: ["12月", "十二月"],
}


@dataclass
class TrainingRecord:
    title: str
    body: str
    category1: str
    category2: str
    link: str
    note_id: str
    pre_interactions: int
    total_interactions: int
    spend: float
    cpe: float
    tokens: set[str] = field(default_factory=set)
    label: bool = False


@dataclass
class Candidate:
    row: dict[str, str]
    publish_date: date
    note_id: str
    title: str
    body: str
    author: str
    link: str
    interactions: int
    likes: int
    collects: int
    comments: int
    shares: int
    category1: str
    category2: str
    sentiment: str
    tokens: set[str]


@dataclass
class Decision:
    method: str
    decision: str
    score: int
    confidence: int
    predicted_cpe: Optional[float] = None
    predicted_cpe_range: str = ""
    suggested_budget: str = ""
    summary: str = ""
    reasons: list[str] = field(default_factory=list)


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.exists() or path.stat().st_size == 0:
        return [], []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), [dict(row) for row in reader]


def load_json(path: Path) -> dict:
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {}


def clean(value: Any) -> str:
    return str(value or "").replace("\u00a0", " ").strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"[\s：:\"'`*（）()/_\\\-.]+", "", clean(value)).lower()


def to_number(value: Any) -> int:
    text = clean(value).replace(",", "").replace("，", "").replace("￥", "").replace("¥", "").replace("元", "")
    if not text or re.fullmatch(r"[-—/\\]+", text):
        return 0
    for suffix, factor in (("万", 10000), ("千", 1000), ("k", 1000), ("K", 1000), ("w", 10000), ("W", 10000)):
        if text.endswith(suffix):
            try:
                return int(float(text[: -len(suffix)]) * factor)
            except Exception:
                return 0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return 0
    try:
        return int(float(match.group(0)))
    except Exception:
        return 0


def to_float(value: Any) -> float:
    text = clean(value).replace(",", "").replace("，", "").replace("￥", "").replace("¥", "").replace("元", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return float("nan")
    try:
        return float(match.group(0))
    except Exception:
        return float("nan")


def parse_date_value(value: Any) -> Optional[date]:
    text = clean(value)
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        n = float(text)
        if n > 10**12:
            return datetime.fromtimestamp(n / 1000).date()
        if n > 10**9:
            return datetime.fromtimestamp(n).date()
        if 30000 <= n <= 60000:
            return date.fromordinal(date(1899, 12, 30).toordinal() + int(n))
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except Exception:
            pass
    match = re.search(r"(20\d{2})[年/\-.](\d{1,2})[月/\-.](\d{1,2})", text)
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return None


def note_id_from_url(value: str) -> str:
    match = re.search(r"/(?:discovery/item|explore|search_result)/([0-9a-zA-Z]{24})", value or "")
    return match.group(1) if match else ""


def tokenize(title: str, body: str = "") -> set[str]:
    text = f"{title}\n{body}".lower()
    text = re.sub(r"https?://\S+", " ", text)
    tokens: set[str] = set()
    for keyword in KEYWORDS:
        if keyword.lower() in text:
            tokens.add(keyword)
    for segment in re.findall(r"[\u4e00-\u9fff]{2,}", text):
        chars = [char for char in segment if char not in STOP_CHARS]
        compact = "".join(chars)
        if 2 <= len(compact) <= 8:
            tokens.add(compact)
        for size in (2, 3):
            for index in range(0, max(0, len(compact) - size + 1)):
                token = compact[index : index + size]
                if not re.fullmatch(r"[的一是在不了和有就人都]+", token):
                    tokens.add(token)
    for token in re.findall(r"[a-z0-9]{2,}", text):
        tokens.add(token)
    return set(list(tokens)[:80])


def jaccard(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / max(1, len(a | b))


def quantile(values: list[float], q: float) -> float:
    clean_values = sorted(v for v in values if math.isfinite(v))
    if not clean_values:
        return 1.4
    if len(clean_values) == 1:
        return clean_values[0]
    pos = (len(clean_values) - 1) * q
    lo = math.floor(pos)
    hi = math.ceil(pos)
    if lo == hi:
        return clean_values[lo]
    return clean_values[lo] * (hi - pos) + clean_values[hi] * (pos - lo)


def sigmoid(value: float) -> float:
    return 1 / (1 + math.exp(-max(-40, min(40, value))))


def interaction_similarity(a: int, b: int) -> float:
    return 1 / (1 + abs(math.log1p(max(0, a)) - math.log1p(max(0, b))))


def parse_history(path: Path) -> list[TrainingRecord]:
    _, rows = read_csv(path)
    records: list[TrainingRecord] = []
    for row in rows:
        title = clean(row.get("标题") or row.get("笔记标题"))
        body = clean(row.get("正文") or row.get("笔记内容") or row.get("内容"))
        spend = to_float(row.get("总投放金额") or row.get("投放金额") or row.get("消耗金额"))
        total_interactions = to_number(row.get("总互动量") or row.get("投后互动量") or row.get("投后笔记总互动量"))
        cpe = to_float(row.get("综合CPE") or row.get("综合cpe") or row.get("CPE"))
        if not math.isfinite(cpe) and spend > 0 and total_interactions > 0:
            cpe = spend / total_interactions
        if not title or not math.isfinite(cpe) or cpe <= 0:
            continue
        pre = to_number(row.get("投前互动量") or row.get("当前互动量"))
        record = TrainingRecord(
            title=title,
            body=body,
            category1=clean(row.get("一级分类") or row.get("内容类型")),
            category2=clean(row.get("二级分类") or row.get("具体产品/场景")),
            link=clean(row.get("链接") or row.get("笔记链接")),
            note_id=clean(row.get("笔记ID") or row.get("note_id")),
            pre_interactions=pre,
            total_interactions=total_interactions,
            spend=spend if math.isfinite(spend) else 0,
            cpe=cpe,
            tokens=tokenize(title, body),
        )
        records.append(record)
    if len(records) < 4:
        raise RuntimeError(f"Hype 历史训练数据不足，无法训练本地加热模型: {path}")
    threshold = quantile([record.cpe for record in records], 0.45)
    for record in records:
        record.label = record.cpe <= threshold
    return records


def build_candidates(rows: list[dict[str, str]], start: date, end: date) -> tuple[list[Candidate], dict[str, int]]:
    candidates: list[Candidate] = []
    seen: set[str] = set()
    total_in_range = 0
    skipped_non_positive = 0
    for row in rows:
        publish_date = parse_date_value(row.get("发布时间") or row.get("发布日期"))
        if not publish_date or publish_date < start or publish_date > end:
            continue
        total_in_range += 1
        excluded = clean(row.get("是否剔除"))
        if excluded and excluded not in ("否", "0", "false", "False"):
            continue
        sentiment = clean(row.get("正负向"))
        if sentiment != "正向":
            skipped_non_positive += 1
            continue
        link = clean(row.get("笔记链接") or row.get("链接"))
        note_id = clean(row.get("笔记ID") or row.get("note_id")) or note_id_from_url(link)
        title = clean(row.get("笔记标题") or row.get("标题"))
        body = clean(row.get("笔记内容") or row.get("内容"))
        if not note_id or not title:
            continue
        key = note_id or link
        if key in seen:
            continue
        seen.add(key)
        likes = to_number(row.get("点赞量"))
        collects = to_number(row.get("收藏量"))
        comments = to_number(row.get("评论量"))
        shares = to_number(row.get("分享量"))
        interactions = to_number(row.get("互动量")) or likes + collects + comments + shares
        candidates.append(
            Candidate(
                row=row,
                publish_date=publish_date,
                note_id=note_id,
                title=title,
                body=body,
                author=clean(row.get("博主昵称") or row.get("作者昵称")),
                link=link,
                interactions=interactions,
                likes=likes,
                collects=collects,
                comments=comments,
                shares=shares,
                category1=clean(row.get("内容类型") or row.get("一级分类")),
                category2=clean(row.get("具体产品/场景") or row.get("二级分类")),
                sentiment=sentiment,
                tokens=tokenize(title, body),
            )
        )
    candidates.sort(key=lambda item: (item.publish_date, item.interactions), reverse=True)
    return candidates, {
        "totalInRange": total_in_range,
        "positiveCandidates": len(candidates),
        "skippedNonPositive": skipped_non_positive,
    }


class HypeLocalPredictor:
    def __init__(self, records: list[TrainingRecord]):
        self.records = records
        self.cpe_threshold = quantile([record.cpe for record in records], 0.45)
        self.median_spend = quantile([record.spend for record in records if record.spend > 0], 0.5)
        self.baseline = sum(1 for record in records if record.label) / len(records)

    def predict(self, candidate: Candidate) -> Decision:
        scored = []
        for record in self.records:
            token_score = jaccard(candidate.tokens, record.tokens)
            category_score = 0.0
            if candidate.category1 and candidate.category1 == record.category1:
                category_score += 0.18
            if candidate.category2 and candidate.category2 == record.category2:
                category_score += 0.22
            interaction_score = interaction_similarity(candidate.interactions, record.pre_interactions)
            score = 0.52 * token_score + category_score + 0.30 * interaction_score
            if score > 0:
                scored.append((score, record))
        scored.sort(key=lambda item: item[0], reverse=True)
        neighbors = scored[:8]
        if not neighbors:
            predicted_cpe = self.cpe_threshold
            success_rate = self.baseline
            confidence = 30
        else:
            weights = [max(0.03, score) * (0.88**index) for index, (score, _) in enumerate(neighbors)]
            weight_sum = sum(weights) or 1
            predicted_cpe = sum(record.cpe * weight for weight, (_, record) in zip(weights, neighbors)) / weight_sum
            success_rate = sum((1 if record.label else 0) * weight for weight, (_, record) in zip(weights, neighbors)) / weight_sum
            confidence = int(max(24, min(92, 28 + sum(score for score, _ in neighbors[:5]) * 52)))
        cpe_probability = sigmoid((self.cpe_threshold - predicted_cpe) / max(0.25, self.cpe_threshold * 0.42))
        interaction_boost = 0.04 if candidate.interactions >= 100 else 0.0
        probability = max(0.02, min(0.98, 0.58 * success_rate + 0.34 * cpe_probability + 0.08 * self.baseline + interaction_boost))
        score = int(round(probability * 100))
        if score >= 70:
            decision = "值得加热"
        elif score >= 55:
            decision = "建议小额测试"
        else:
            decision = "暂不建议加热"
        reasons = []
        for neighbor_score, record in neighbors[:3]:
            reasons.append(f"相似{round(neighbor_score * 100)}%：{record.title[:28]}，CPE {record.cpe:.2f}")
        return Decision(
            method="hype",
            decision=decision,
            score=score,
            confidence=confidence,
            predicted_cpe=predicted_cpe,
            predicted_cpe_range=f"{max(0.05, predicted_cpe * 0.82):.2f}-{predicted_cpe * 1.18:.2f}",
            suggested_budget=self.suggest_budget(decision, predicted_cpe),
            summary=f"Hype历史模型预测CPE {predicted_cpe:.2f}，优秀线 {self.cpe_threshold:.2f}，相似样本达标率 {success_rate:.0%}。",
            reasons=reasons,
        )

    def suggest_budget(self, decision: str, predicted_cpe: float) -> str:
        median_spend = self.median_spend if self.median_spend > 0 else 100
        ratio = self.cpe_threshold / predicted_cpe if predicted_cpe > 0 else 1
        if decision == "值得加热":
            base = max(60, min(260, median_spend * (0.32 if ratio > 1.2 else 0.24)))
            return f"{round(base / 10) * 10}-{round(base * 1.8 / 10) * 10}元"
        if decision == "建议小额测试":
            base = max(30, min(120, median_spend * 0.16))
            return f"{round(base / 10) * 10}-{round(base * 1.6 / 10) * 10}元"
        return "0元"


def resolve_llm_config(args: argparse.Namespace) -> dict[str, str]:
    config = load_json(HYPE_LLM_CONFIG)
    return {
        "api_key": args.api_key or os.environ.get("OPENAI_API_KEY") or config.get("apiKey") or "",
        "base_url": args.base_url or os.environ.get("OPENAI_BASE_URL") or config.get("baseUrl") or "https://llm-proxy.intra.xiaojukeji.com",
        "model": args.model or config.get("model") or "kimi-k2.5-external",
    }


def endpoint_candidates(base_url: str) -> list[str]:
    base = base_url.rstrip("/")
    urls = [f"{base}/chat/completions"]
    if not base.endswith("/v1"):
        urls.append(f"{base}/v1/chat/completions")
    return urls


class LlmHttpError(RuntimeError):
    def __init__(self, status_code: int, body: str):
        super().__init__(f"HTTP {status_code}: {body[:500]}")
        self.status_code = status_code


def curl_config_quote(value: str) -> str:
    return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'


def curl_post_json(endpoint: str, body: bytes, api_key: str, timeout: float) -> dict:
    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile("wb", delete=False) as handle:
            handle.write(body)
            temp_path = handle.name
        config = "\n".join(
            [
                f"url = {curl_config_quote(endpoint)}",
                'request = "POST"',
                'header = "Content-Type: application/json"',
                f"header = {curl_config_quote('Authorization: Bearer ' + api_key)}",
                f"data-binary = @{temp_path}",
            ]
        )
        result = subprocess.run(
            [
                "curl",
                "-sS",
                "--http1.1",
                "--tlsv1.2",
                "--no-keepalive",
                "--retry",
                "3",
                "--retry-delay",
                "2",
                "--retry-max-time",
                str(max(30, int(timeout) + 25)),
                "--retry-all-errors",
                "--connect-timeout",
                "30",
                "--max-time",
                str(max(30, int(timeout) + 10)),
                "-w",
                "\n%{http_code}",
                "--config",
                "-",
            ],
            input=config,
            text=True,
            capture_output=True,
            check=False,
        )
        if result.returncode != 0:
            raise RuntimeError((result.stderr or result.stdout or f"curl exit {result.returncode}").strip())
        response_text, _, status_text = result.stdout.rpartition("\n")
        status = int(status_text) if status_text.isdigit() else 0
        if status < 200 or status >= 300:
            raise LlmHttpError(status, response_text)
        return json.loads(response_text)
    finally:
        if temp_path:
            try:
                os.unlink(temp_path)
            except OSError:
                pass


def urllib_post_json(endpoint: str, body: bytes, api_key: str, timeout: float) -> dict:
    request = urllib.request.Request(
        endpoint,
        data=body,
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}", "Connection": "close"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raise LlmHttpError(exc.code, exc.read().decode("utf-8", errors="replace")) from exc


def post_json_with_fallback(endpoint: str, body: bytes, api_key: str, timeout: float) -> dict:
    try:
        return urllib_post_json(endpoint, body, api_key, min(timeout, 20))
    except LlmHttpError:
        raise
    except Exception as first_error:
        try:
            return curl_post_json(endpoint, body, api_key, timeout)
        except Exception as second_error:
            first = re.sub(r"\s+", " ", str(first_error)).strip()
            second = re.sub(r"\s+", " ", str(second_error)).strip()
            raise RuntimeError(f"urllib失败: {first}; curl兜底失败: {second}") from second_error


def parse_json_from_text(text: str) -> dict:
    clean_text = clean(text)
    clean_text = re.sub(r"^```(?:json)?\s*", "", clean_text)
    clean_text = re.sub(r"\s*```$", "", clean_text)
    try:
        return json.loads(clean_text)
    except Exception:
        match = re.search(r"\{.*\}", clean_text, flags=re.S)
        if match:
            return json.loads(match.group(0))
        raise


def normalize_ai_decision(payload: dict[str, Any]) -> Decision:
    decision = clean(payload.get("decision"))
    if "值得加热" in decision and "不" not in decision:
        decision = "值得加热"
    elif "小额" in decision or "测试" in decision:
        decision = "建议小额测试"
    elif "不建议" in decision or "暂不" in decision:
        decision = "暂不建议加热"
    else:
        decision = "暂不建议加热"
    reasons = payload.get("reasons")
    if not isinstance(reasons, list):
        reasons = []
    return Decision(
        method="ai",
        decision=decision,
        score=max(0, min(100, int(to_number(payload.get("score"))))),
        confidence=max(0, min(100, int(to_number(payload.get("confidence"))))),
        predicted_cpe_range=clean(payload.get("predicted_cpe_range") or payload.get("predictedCpeRange")),
        suggested_budget=clean(payload.get("suggested_budget") or payload.get("suggestedBudget")),
        summary=clean(payload.get("summary")),
        reasons=[clean(item) for item in reasons[:5] if clean(item)],
    )


def build_ai_prompt(candidate: Candidate, local_decision: Decision) -> list[dict[str, str]]:
    system = "你是小红书口碑加热投放判断助手。必须以历史CPE、相似样本和当前互动为主要依据，输出严格JSON。"
    user = f"""
请判断这条小红书笔记是否值得进入加热投放候选池。

决策只能三选一：
- 值得加热：预期CPE好、相似历史表现较好，适合进入Excel候选表
- 建议小额测试：有潜力但不稳定，可以小预算测试
- 暂不建议加热：风险较高或与历史高CPE样本相似

只返回JSON，不要Markdown：
{{
  "decision": "值得加热 | 建议小额测试 | 暂不建议加热",
  "score": 0,
  "confidence": 0,
  "predicted_cpe_range": "",
  "suggested_budget": "",
  "summary": "",
  "reasons": []
}}

笔记信息：
{json.dumps({
    "发布时间": candidate.publish_date.isoformat(),
    "标题": candidate.title,
    "正文": candidate.body[:1600],
    "作者昵称": candidate.author,
    "笔记ID": candidate.note_id,
    "投前互动量": candidate.interactions,
    "点赞量": candidate.likes,
    "收藏量": candidate.collects,
    "评论量": candidate.comments,
    "分享量": candidate.shares,
    "一级分类": candidate.category1,
    "二级分类": candidate.category2,
    "Hype本地判断": local_decision.decision,
    "Hype值得分": local_decision.score,
    "Hype预测CPE": local_decision.predicted_cpe,
    "Hype理由": local_decision.reasons,
}, ensure_ascii=False, indent=2)}
""".strip()
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def call_llm(messages: list[dict[str, str]], model: str, api_key: str, base_url: str, timeout: float, retries: int = 3) -> Decision:
    body = json.dumps(
        {
            "model": model,
            "messages": messages,
            "temperature": 0.15,
            "response_format": {"type": "json_object"},
        },
        ensure_ascii=False,
    ).encode("utf-8")
    last_error: Optional[Exception] = None
    for endpoint in endpoint_candidates(base_url):
        for attempt in range(1, max(1, retries) + 1):
            try:
                payload = post_json_with_fallback(endpoint, body, api_key, timeout)
                content = payload["choices"][0]["message"]["content"]
                return normalize_ai_decision(parse_json_from_text(content))
            except LlmHttpError as exc:
                last_error = exc
                if exc.status_code == 404:
                    break
                if exc.status_code in (408, 409, 425, 429, 500, 502, 503, 504) and attempt < retries:
                    time.sleep(min(8, 1.5 * attempt))
                    continue
                break
            except Exception as exc:
                last_error = exc
                if attempt < retries:
                    time.sleep(min(8, 1.5 * attempt))
                    continue
                break
    raise RuntimeError(str(last_error) if last_error else "LLM request failed")


def ai_predict(candidate: Candidate, local_decision: Decision, config: dict[str, str], timeout: float, retries: int) -> Decision:
    if not config["api_key"]:
        raise RuntimeError("未配置 LLM API Key")
    models = MODEL_GROUPS.get(config["model"], [config["model"]])
    last_error = ""
    for model in models:
        try:
            decision = call_llm(build_ai_prompt(candidate, local_decision), model, config["api_key"], config["base_url"], timeout, retries)
            decision.method = "ai"
            return decision
        except Exception as exc:
            last_error = f"{model}: {exc}"
    raise RuntimeError(last_error or "AI 判断失败")


def ai_error_fallback(local_decision: Decision, error: Exception) -> Decision:
    decision = copy(local_decision)
    decision.method = "hype-fallback"
    reason = re.sub(r"\s+", " ", str(error)).strip()
    summary = decision.summary or ""
    decision.summary = f"AI连接失败，使用Hype本地模型兜底。{summary}".strip()
    if reason:
        decision.reasons = [f"AI连接失败：{reason[:180]}"] + decision.reasons[:4]
    return decision


def decision_passes(decision: Decision, min_decision: str) -> bool:
    if min_decision == "test":
        return decision.decision in ("值得加热", "建议小额测试")
    return decision.decision == "值得加热"


def require_openpyxl():
    attempted_paths: list[str] = []
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception as first_exc:
        candidate_paths: list[Path] = []
        try:
            candidate_paths.append(Path(site.getusersitepackages()))
        except Exception:
            pass
        try:
            candidate_paths.extend(Path(item) for item in site.getsitepackages())
        except Exception:
            pass
        candidate_paths.extend(Path.home().glob("Library/Python/*/lib/python/site-packages"))
        candidate_paths.extend(Path.home().glob(".local/lib/python*/site-packages"))

        for candidate in candidate_paths:
            if not candidate.exists():
                continue
            text = str(candidate)
            attempted_paths.append(text)
            if text not in sys.path:
                sys.path.insert(0, text)
            try:
                importlib.invalidate_caches()
                import openpyxl  # type: ignore
                return openpyxl
            except Exception:
                continue

        detail = (
            "缺少 openpyxl，无法写入 xlsx。\n"
            f"当前 Python: {sys.executable}\n"
            f"Python 版本: {sys.version.split()[0]}\n"
            f"首次 import 错误: {first_exc}\n"
            f"已尝试 site-packages: {attempted_paths or '无'}\n"
            "请用同一个 Python 安装："
            f"{sys.executable} -m pip install --user openpyxl"
        )
        raise RuntimeError(detail) from first_exc


def header_map(ws: Any) -> dict[str, int]:
    raw = {normalize_header(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    mapped: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            col = raw.get(normalize_header(alias))
            if col:
                mapped[key] = col
                break
    return mapped


def sheet_for_month(wb: Any, month: int) -> Any:
    names = MONTH_NAMES.get(month, [f"{month}月"])
    for ws in wb.worksheets:
        title = ws.title
        if any(name in title for name in names) and "口碑加热" in title and "废表" not in title:
            return ws
    template = wb["6月口碑加热"] if "6月口碑加热" in wb.sheetnames else wb.worksheets[-1]
    ws = wb.copy_worksheet(template)
    ws.title = f"{month}月口碑加热"
    for row in range(3, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None
    return ws


def existing_note_keys(wb: Any) -> set[str]:
    keys: set[str] = set()
    for ws in wb.worksheets:
        mapping = header_map(ws)
        note_col = mapping.get("note_id")
        link_col = mapping.get("link")
        if not note_col and not link_col:
            continue
        for row in range(3, ws.max_row + 1):
            if note_col:
                value = clean(ws.cell(row, note_col).value)
                if value:
                    keys.add(value)
            if link_col:
                value = clean(ws.cell(row, link_col).value)
                note_id = note_id_from_url(value)
                if note_id:
                    keys.add(note_id)
                elif value:
                    keys.add(value)
    return keys


def first_empty_row(ws: Any, mapping: dict[str, int]) -> int:
    key_cols = [mapping.get("note_id"), mapping.get("link"), mapping.get("title")]
    key_cols = [col for col in key_cols if col]
    for row in range(3, max(ws.max_row, 3) + 1):
        if all(not clean(ws.cell(row, col).value) for col in key_cols):
            return row
    return ws.max_row + 1


def copy_row_style(ws: Any, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)


def set_if_present(ws: Any, mapping: dict[str, int], row: int, key: str, value: Any) -> None:
    col = mapping.get(key)
    if col:
        ws.cell(row, col).value = value


def col_letter(col: Optional[int]) -> str:
    if not col:
        return ""
    try:
        from openpyxl.utils import get_column_letter  # type: ignore

        return get_column_letter(col)
    except Exception:
        result = ""
        value = col
        while value:
            value, rem = divmod(value - 1, 26)
            result = chr(65 + rem) + result
        return result


def append_to_workbook(workbook_path: Path, selected: list[tuple[Candidate, Decision]], dry_run: bool) -> dict[str, Any]:
    if not selected:
        return {"appended": 0, "skippedExisting": 0, "sheets": {}, "excelWriteDetails": []}
    openpyxl = require_openpyxl()
    wb = openpyxl.load_workbook(workbook_path)
    existing = existing_note_keys(wb)
    appended = 0
    skipped = 0
    sheets: dict[str, int] = {}
    excel_details: list[str] = []
    for candidate, decision in selected:
        key = candidate.note_id or candidate.link
        if key in existing:
            skipped += 1
            continue
        ws = sheet_for_month(wb, candidate.publish_date.month)
        mapping = header_map(ws)
        row_index = first_empty_row(ws, mapping)
        if row_index > ws.max_row:
            copy_row_style(ws, max(3, ws.max_row), row_index)
        remark = f"{decision.method.upper()}判断：{decision.decision}，值得分{decision.score}，置信度{decision.confidence}。"
        if decision.predicted_cpe:
            remark += f" 预测CPE {decision.predicted_cpe:.2f}。"
        if decision.summary:
            remark += f" {decision.summary}"
        set_if_present(ws, mapping, row_index, "publish_date", candidate.publish_date)
        publish_col = col_letter(mapping.get("publish_date"))
        end_value = f"={publish_col}{row_index}+7" if publish_col else None
        set_if_present(ws, mapping, row_index, "end_date", end_value)
        set_if_present(ws, mapping, row_index, "status", "待评估")
        set_if_present(ws, mapping, row_index, "author", candidate.author)
        set_if_present(ws, mapping, row_index, "link", candidate.link)
        set_if_present(ws, mapping, row_index, "note_id", candidate.note_id)
        set_if_present(ws, mapping, row_index, "ad_records", "")
        set_if_present(ws, mapping, row_index, "title", candidate.title)
        set_if_present(ws, mapping, row_index, "pre_interactions", candidate.interactions)
        set_if_present(ws, mapping, row_index, "post_likes", "")
        set_if_present(ws, mapping, row_index, "post_collects", "")
        set_if_present(ws, mapping, row_index, "post_comments", "")
        set_if_present(ws, mapping, row_index, "post_total_interactions", "")
        set_if_present(ws, mapping, row_index, "post_increment_interactions", "")
        set_if_present(ws, mapping, row_index, "total_spend", "")
        spend_col = col_letter(mapping.get("total_spend"))
        increment_col = col_letter(mapping.get("post_increment_interactions"))
        post_total_col = col_letter(mapping.get("post_total_interactions"))
        if spend_col and increment_col:
            set_if_present(ws, mapping, row_index, "placement_cpe", f'=IFERROR({spend_col}{row_index}/{increment_col}{row_index},"")')
        else:
            set_if_present(ws, mapping, row_index, "placement_cpe", "")
        if spend_col and post_total_col:
            set_if_present(ws, mapping, row_index, "combined_cpe", f'=IFERROR({spend_col}{row_index}/{post_total_col}{row_index},"")')
        elif spend_col and increment_col:
            set_if_present(ws, mapping, row_index, "combined_cpe", f'=IFERROR({spend_col}{row_index}/{increment_col}{row_index},"")')
        else:
            set_if_present(ws, mapping, row_index, "combined_cpe", "")
        set_if_present(ws, mapping, row_index, "category1", candidate.category1 or "待分类")
        set_if_present(ws, mapping, row_index, "category2", candidate.category2 or "待分类")
        set_if_present(ws, mapping, row_index, "remark", remark[:32000])
        if post_total_col:
            set_if_present(
                ws,
                mapping,
                row_index,
                "viral_level",
                f'=IF({post_total_col}{row_index}="","",IF({post_total_col}{row_index}>=5000,"S 级",IF({post_total_col}{row_index}>=1000,"A 级",IF({post_total_col}{row_index}>=500,"B 级","C 级"))))',
            )
        else:
            set_if_present(ws, mapping, row_index, "viral_level", "高潜" if decision.score >= 85 else "潜力")
        if len(excel_details) < 20:
            detail_fields = {
                "sheet": ws.title,
                "row": row_index,
                "发布日期": candidate.publish_date.isoformat(),
                "投放截止": end_value,
                "状态": "待评估",
                "作者昵称": candidate.author,
                "链接": candidate.link,
                "笔记ID": candidate.note_id,
                "标题": candidate.title,
                "投前互动量": candidate.interactions,
                "一级分类": candidate.category1 or "待分类",
                "二级分类": candidate.category2 or "待分类",
                "备注": remark[:260],
            }
            excel_details.append(
                f"{'预写入' if dry_run else '写入'}Excel："
                + "；".join(f"{field}={clean(value)}" for field, value in detail_fields.items())
            )
        existing.add(key)
        appended += 1
        sheets[ws.title] = sheets.get(ws.title, 0) + 1
    if not dry_run and appended:
        wb.save(workbook_path)
    return {
        "appended": appended if not dry_run else 0,
        "wouldAppend": appended,
        "skippedExisting": skipped,
        "sheets": sheets,
        "excelWriteDetails": excel_details,
        "excelWriteDetailsOmitted": max(0, appended - len(excel_details)),
    }


def preview_item(candidate: Candidate, decision: Decision) -> dict[str, Any]:
    return {
        "发布时间": candidate.publish_date.isoformat(),
        "标题": candidate.title,
        "笔记ID": candidate.note_id,
        "互动量": candidate.interactions,
        "作者": candidate.author,
        "正负向": candidate.sentiment,
        "判断": decision.decision,
        "值得分": decision.score,
        "置信度": decision.confidence,
        "预测CPE": round(decision.predicted_cpe, 2) if decision.predicted_cpe else decision.predicted_cpe_range,
    }


def run(args: argparse.Namespace) -> dict[str, Any]:
    start = parse_date_value(args.start_date)
    end = parse_date_value(args.end_date)
    if not start or not end:
        raise RuntimeError("请选择有效的开始日期和结束日期")
    if end < start:
        raise RuntimeError("结束日期不能早于开始日期")

    _, table_rows = read_csv(Path(args.source))
    if not table_rows:
        raise RuntimeError(f"总表为空或不存在: {args.source}")
    candidates, candidate_stats = build_candidates(table_rows, start, end)
    positive_candidates = len(candidates)
    if args.limit > 0:
        candidates = candidates[: args.limit]

    history = parse_history(Path(args.history))
    predictor = HypeLocalPredictor(history)
    config = resolve_llm_config(args)

    selected: list[tuple[Candidate, Decision]] = []
    judged = 0
    ai_fallback_to_hype = 0
    errors: list[str] = []
    for index, candidate in enumerate(candidates, start=1):
        local_decision = predictor.predict(candidate)
        decision = local_decision
        if args.method == "ai":
            try:
                decision = ai_predict(candidate, local_decision, config, args.timeout, args.retries)
            except Exception as exc:
                decision = ai_error_fallback(local_decision, exc)
                ai_fallback_to_hype += 1
            if args.delay > 0 and index < len(candidates):
                time.sleep(args.delay)
        elif args.method == "both":
            try:
                ai_decision = ai_predict(candidate, local_decision, config, args.timeout, args.retries)
                decision = ai_decision
                if not decision_passes(local_decision, args.min_decision):
                    decision.decision = "暂不建议加热"
                    decision.summary = "AI入选但Hype本地模型未达入选门槛，已按交集策略剔除。"
            except Exception as exc:
                decision = ai_error_fallback(local_decision, exc)
                ai_fallback_to_hype += 1
            if args.delay > 0 and index < len(candidates):
                time.sleep(args.delay)
        judged += 1
        if decision_passes(decision, args.min_decision):
            selected.append((candidate, decision))

    workbook_result = append_to_workbook(Path(args.workbook), selected, args.dry_run)
    return {
        "source": str(Path(args.source)),
        "history": str(Path(args.history)),
        "workbook": str(Path(args.workbook)),
        "method": args.method,
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "totalInRange": candidate_stats["totalInRange"],
        "positiveCandidates": positive_candidates,
        "skippedNonPositive": candidate_stats["skippedNonPositive"],
        "judged": judged,
        "aiFallbackToHype": ai_fallback_to_hype,
        "selected": len(selected),
        "dryRun": args.dry_run,
        **workbook_result,
        "errors": errors[:8],
        "preview": [preview_item(candidate, decision) for candidate, decision in selected[:10]],
    }


def main(argv: Optional[Iterable[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Select XHS amplification candidates and append them to Hype workbook.")
    parser.add_argument("--source", default=str(DATA_TABLE), help="Pipeline monitoring data table CSV.")
    parser.add_argument("--history", default=str(HYPE_HISTORY), help="Hype_Something historical training CSV.")
    parser.add_argument("--workbook", default=str(HYPE_WORKBOOK), help="Target amplification workbook.")
    parser.add_argument("--start-date", required=True, help="Inclusive start date, e.g. 2026-06-01.")
    parser.add_argument("--end-date", required=True, help="Inclusive end date, e.g. 2026-06-30.")
    parser.add_argument("--method", choices=["hype", "ai", "both"], default="hype")
    parser.add_argument("--min-decision", choices=["worth", "test"], default="worth", help="worth=only 值得加热, test=also 建议小额测试.")
    parser.add_argument("--limit", type=int, default=30, help="Maximum rows to judge after date filtering. 0 means all.")
    parser.add_argument("--dry-run", action="store_true", help="Preview only, do not write workbook.")
    parser.add_argument("--model", choices=MODEL_OPTIONS, default="", help="LLM model option.")
    parser.add_argument("--api-key", default="", help="LLM API key. Defaults to env/Hype_Something config.")
    parser.add_argument("--base-url", default="", help="LLM base URL. Defaults to env/Hype_Something config.")
    parser.add_argument("--delay", type=float, default=0.8, help="Seconds between AI calls.")
    parser.add_argument("--timeout", type=float, default=90.0)
    parser.add_argument("--retries", type=int, default=3, help="Retry count per candidate/model for transient LLM transport errors.")
    args = parser.parse_args(argv)

    try:
        result = run(args)
    except Exception as exc:
        print(f"ERROR: {exc}", flush=True)
        return 1
    print(json.dumps(result, ensure_ascii=False, indent=2), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
