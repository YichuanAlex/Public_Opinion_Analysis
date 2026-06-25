#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按监控总表逐帖导出评论区数据到 Comment_Data/{xhs,dy}/{笔记ID}.xlsx。

默认保护：
  - 默认处理监控总表中全部候选帖子；
  - 默认导出每条帖子的全部评论；
  - 已存在对应 xlsx 文件则跳过；
  - 帖间等待 8 秒；
  - 支持 --dry-run 预览，不访问平台。

示例：
  python3 Pipeline/export_comment_sections.py --dry-run
  python3 Pipeline/export_comment_sections.py --platform all --reset-output --overwrite --headed
  python3 Pipeline/export_comment_sections.py --platform xhs --reset-output --overwrite --headed
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import Any


PIPELINE_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = PIPELINE_DIR.parent
COMMENT_ROOT = PROJECT_ROOT / "Comment_Data"
TMP_DIR = COMMENT_ROOT / ".tmp"


def candidate_pythons() -> list[str]:
    raw = [
        os.environ.get("PIPELINE_PYTHON", ""),
        str(PROJECT_ROOT / ".venv" / "bin" / "python"),
        "/Library/Developer/CommandLineTools/usr/bin/python3",
        "/usr/bin/python3",
        sys.executable,
        "python3",
    ]
    out: list[str] = []
    seen: set[str] = set()
    for item in raw:
        if not item:
            continue
        if item in seen:
            continue
        if "/" in item and not Path(item).exists():
            continue
        out.append(item)
        seen.add(item)
    return out


def can_import(python_bin: str, module: str) -> bool:
    try:
        result = subprocess.run(
            [python_bin, "-c", f"import {module}"],
            cwd=str(PROJECT_ROOT),
            text=True,
            capture_output=True,
            timeout=10,
            check=False,
        )
        return result.returncode == 0
    except Exception:
        return False


def reexec_with_openpyxl_if_needed() -> None:
    try:
        import openpyxl  # noqa: F401
        return
    except Exception:
        pass
    for python_bin in candidate_pythons():
        if Path(python_bin).resolve() == Path(sys.executable).resolve():
            continue
        if can_import(python_bin, "openpyxl"):
            os.execv(python_bin, [python_bin] + sys.argv)
    raise RuntimeError("缺少 openpyxl，无法写入 xlsx。请先安装到当前 Python，或使用项目启动脚本选择的 Python。")


reexec_with_openpyxl_if_needed()

import openpyxl  # type: ignore  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border  # type: ignore  # noqa: E402
from openpyxl.utils import get_column_letter  # type: ignore  # noqa: E402


PLATFORMS = {
    "xhs": {
        "name": "小红书",
        "table": PIPELINE_DIR / "xhs_Data_Table_on_Channel_Public_Opinion_Monitoring_2026.csv",
        "origin": PIPELINE_DIR / "xhs_origin_data.csv",
        "script": PIPELINE_DIR / "xhs_comment_to_csv.py",
        "out_dir": COMMENT_ROOT / "xhs",
        "id_pattern": re.compile(r"(?:/explore/|/discovery/item/|/search_result/)([0-9a-zA-Z]{24})"),
        "default_interval": 1.2,
    },
    "dy": {
        "name": "抖音",
        "table": PIPELINE_DIR / "dy_Data_Table_on_Channel_Public_Opinion_Monitoring_2026.csv",
        "origin": PIPELINE_DIR / "dy_origin_data.csv",
        "script": PIPELINE_DIR / "dy_comment_to_csv.py",
        "out_dir": COMMENT_ROOT / "dy",
        "id_pattern": re.compile(r"(?:modal_id=|/video/|/note/|share/video/)(\d{10,30})"),
        "default_interval": 1.5,
    },
}

TEMPLATE_CANDIDATES = [
    COMMENT_ROOT / "demo" / "雨林大巴评论区.xlsx",
    COMMENT_ROOT / "others" / "雨林大巴评论区.xlsx",
    PIPELINE_DIR / "xhs_Comment_Section_Data_Dictionary.xlsx",
]

OUTPUT_HEADERS = [
    "评论ID",
    "评论内容",
    "评论图片链接",
    "点赞量",
    "评论时间",
    "IP地址",
    "子评论数",
    "笔记ID",
    "笔记链接",
    "用户ID",
    "用户链接",
    "用户名称",
    "一级评论ID",
    "一级评论内容",
    "评论层级",
    "回复目标评论ID",
    "回复目标用户ID",
    "回复目标用户名称",
]

FIELD_DESCRIPTIONS = {
    "评论ID": "平台侧评论唯一 ID",
    "评论内容": "评论正文文本",
    "评论图片链接": "评论携带图片链接，如平台返回",
    "点赞量": "评论点赞数",
    "评论时间": "评论发布时间",
    "IP地址": "评论展示 IP 归属地",
    "子评论数": "一级评论下的回复数量",
    "笔记ID": "原帖唯一 ID",
    "笔记链接": "原帖链接",
    "用户ID": "评论用户 ID",
    "用户链接": "评论用户主页链接",
    "用户名称": "评论用户昵称",
    "一级评论ID": "若为子评论，记录所属一级评论 ID",
    "一级评论内容": "若为子评论，记录所属一级评论正文",
    "评论层级": "一级评论/子评论",
    "回复目标评论ID": "回复目标评论 ID",
    "回复目标用户ID": "回复目标用户 ID",
    "回复目标用户名称": "回复目标用户昵称",
}


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def is_blank(value: Any) -> bool:
    return clean_text(value).lower() in {"", "none", "nan", "null", "undefined"}


def safe_filename(value: str) -> str:
    text = re.sub(r"[^\w.\-\u4e00-\u9fff]+", "_", value, flags=re.UNICODE).strip("_")
    return text[:120] or "comment_section"


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def extract_note_id(link: str, pattern: re.Pattern[str]) -> str:
    match = pattern.search(str(link or ""))
    return match.group(1) if match else ""


def origin_link_index(platform: str) -> dict[str, str]:
    config = PLATFORMS[platform]
    rows = read_csv(config["origin"])
    index: dict[str, str] = {}
    for row in rows:
        note_id = clean_text(row.get("note_id") or row.get("aweme_id"))
        link = clean_text(row.get("source_url"))
        if not note_id:
            note_id = extract_note_id(link, config["id_pattern"])
        if not note_id or not link:
            continue
        if platform == "xhs" and "xsec_token=" not in link:
            continue
        index.setdefault(note_id, link)
    return index


def candidate_posts(platform: str, include_zero_comments: bool) -> list[dict[str, str]]:
    config = PLATFORMS[platform]
    rows = read_csv(config["table"])
    links = origin_link_index(platform)
    posts: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in rows:
        note_id = clean_text(row.get("笔记ID")) or extract_note_id(clean_text(row.get("笔记链接")), config["id_pattern"])
        if not note_id or note_id in seen:
            continue
        link = links.get(note_id) or clean_text(row.get("笔记链接"))
        if platform == "xhs" and "xsec_token=" not in link:
            continue
        if not include_zero_comments:
            try:
                if int(float(str(row.get("评论量") or "0").replace(",", ""))) <= 0:
                    continue
            except Exception:
                pass
        seen.add(note_id)
        posts.append({
            "note_id": note_id,
            "url": link,
            "title": clean_text(row.get("笔记标题")),
            "author": clean_text(row.get("博主昵称")),
            "publish_time": clean_text(row.get("发布时间")),
            "comment_count": clean_text(row.get("评论量")),
            "platform": platform,
        })
    return posts


def choose_script_python() -> str:
    for python_bin in candidate_pythons():
        if can_import(python_bin, "openpyxl"):
            return python_bin
    return sys.executable


def run_comment_export(
    platform: str,
    post: dict[str, str],
    csv_path: Path,
    limit_comments: int,
    request_interval: float,
    headed: bool,
    no_sub_comments: bool,
) -> str:
    config = PLATFORMS[platform]
    cmd = [
        choose_script_python(),
        str(config["script"]),
        post["url"],
        "--output",
        str(csv_path),
        "--limit",
        str(max(0, limit_comments)),
        "--request-interval",
        str(request_interval),
        "--use-default-profile",
    ]
    if headed:
        cmd.append("--headed")
    if no_sub_comments:
        cmd.append("--no-sub-comments")
    result = subprocess.run(
        cmd,
        cwd=str(PROJECT_ROOT),
        text=True,
        capture_output=True,
        check=False,
    )
    output = "\n".join(part for part in [result.stdout.strip(), result.stderr.strip()] if part)
    if result.returncode != 0:
        raise RuntimeError(output or f"{config['script'].name} exited with {result.returncode}")
    return output


def find_comment_image(row: dict[str, str]) -> str:
    for key, value in row.items():
        lower = key.lower()
        text = clean_text(value)
        if not text.startswith(("http://", "https://")):
            continue
        if any(token in lower for token in ("image", "picture", "pic", "comment.pictures", "comment.image")):
            return text
    return ""


def rows_for_xlsx(csv_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    root_content: dict[str, str] = {}
    for row in csv_rows:
        if row.get("评论层级") == "一级评论" and row.get("评论ID"):
            root_content[row["评论ID"]] = clean_text(row.get("评论内容"))

    out: list[dict[str, Any]] = []
    for row in csv_rows:
        root_id = clean_text(row.get("一级评论ID"))
        item = {
            "评论ID": row.get("评论ID", ""),
            "评论内容": row.get("评论内容", ""),
            "评论图片链接": find_comment_image(row),
            "点赞量": row.get("点赞量", row.get("点赞数", "")),
            "评论时间": row.get("评论时间", ""),
            "IP地址": row.get("IP地址", ""),
            "子评论数": row.get("子评论数", ""),
            "笔记ID": row.get("笔记ID", ""),
            "笔记链接": row.get("笔记链接", ""),
            "用户ID": row.get("用户ID", ""),
            "用户链接": row.get("用户链接", ""),
            "用户名称": row.get("用户名称", ""),
            "一级评论ID": root_id,
            "一级评论内容": root_content.get(root_id, ""),
            "评论层级": row.get("评论层级", ""),
            "回复目标评论ID": row.get("回复目标评论ID", ""),
            "回复目标用户ID": row.get("回复目标用户ID", ""),
            "回复目标用户名称": row.get("回复目标用户名称", ""),
        }
        out.append(item)
    return out


def apply_sheet_style(ws: Any) -> None:
    header_fill = PatternFill("solid", fgColor="0B5CFF")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "A": 22,
        "B": 46,
        "C": 38,
        "D": 10,
        "E": 20,
        "F": 12,
        "G": 10,
        "H": 24,
        "I": 56,
        "J": 24,
        "K": 42,
        "L": 18,
        "M": 22,
        "N": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for index in range(15, len(OUTPUT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_comment_xlsx(post: dict[str, str], csv_path: Path, xlsx_path: Path) -> int:
    rows = read_csv(csv_path)
    data = rows_for_xlsx(rows)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评论区数据"
    ws.append(OUTPUT_HEADERS)
    for row in data:
        ws.append([row.get(header, "") for header in OUTPUT_HEADERS])
    apply_sheet_style(ws)

    info = wb.create_sheet("帖子信息")
    info_rows = [
        ("平台", "小红书" if post["platform"] == "xhs" else "抖音"),
        ("笔记ID", post["note_id"]),
        ("笔记标题", post.get("title", "")),
        ("博主昵称", post.get("author", "")),
        ("发布时间", post.get("publish_time", "")),
        ("原评论量", post.get("comment_count", "")),
        ("笔记链接", post.get("url", "")),
        ("导出时间", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("实际导出评论数", len(data)),
    ]
    for item in info_rows:
        info.append(item)
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 96
    for cell in info["A"]:
        cell.font = Font(bold=True, color="0B2B66")

    dictionary = wb.create_sheet("字段字典")
    dictionary.append(["字段", "说明"])
    for header in OUTPUT_HEADERS:
        dictionary.append([header, FIELD_DESCRIPTIONS.get(header, "")])
    apply_sheet_style(dictionary)
    dictionary.column_dimensions["A"].width = 22
    dictionary.column_dimensions["B"].width = 72

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    return len(data)


def export_platform(platform: str, args: argparse.Namespace) -> dict[str, Any]:
    config = PLATFORMS[platform]
    config["out_dir"].mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    if args.reset_output and not args.dry_run:
        for old_file in config["out_dir"].glob("*.xlsx"):
            old_file.unlink(missing_ok=True)
    posts = candidate_posts(platform, include_zero_comments=args.include_zero_comments)
    if args.note_id:
        wanted = {item.strip() for item in args.note_id.split(",") if item.strip()}
        posts = [post for post in posts if post["note_id"] in wanted]
    if args.max_posts_per_platform and args.max_posts_per_platform > 0:
        posts = posts[: args.max_posts_per_platform]

    result = {
        "platform": platform,
        "platformName": config["name"],
        "candidatePosts": len(posts),
        "exported": 0,
        "skippedExisting": 0,
        "skippedDryRun": 0,
        "resetOutput": bool(args.reset_output and not args.dry_run),
        "errors": [],
        "files": [],
    }

    for index, post in enumerate(posts, start=1):
        xlsx_path = config["out_dir"] / f"{safe_filename(post['note_id'])}.xlsx"
        if xlsx_path.exists() and not args.overwrite:
            result["skippedExisting"] += 1
            result["files"].append(str(xlsx_path))
            continue
        if args.dry_run:
            result["skippedDryRun"] += 1
            result["files"].append(str(xlsx_path))
            continue
        csv_path = TMP_DIR / f"{platform}_{safe_filename(post['note_id'])}_{int(time.time())}.csv"
        try:
            interval = args.request_interval if args.request_interval is not None else config["default_interval"]
            stdout = run_comment_export(
                platform,
                post,
                csv_path,
                limit_comments=args.limit_comments,
                request_interval=interval,
                headed=args.headed,
                no_sub_comments=args.no_sub_comments,
            )
            count = write_comment_xlsx(post, csv_path, xlsx_path)
            result["exported"] += 1
            result["files"].append(str(xlsx_path))
            print(f"[{config['name']}] {index}/{len(posts)} {post['note_id']} 导出 {count} 条评论 -> {xlsx_path}")
            if stdout:
                print(stdout)
        except Exception as exc:
            result["errors"].append({"note_id": post["note_id"], "error": str(exc)})
            print(f"[{config['name']}] ERROR {post['note_id']}: {exc}")
        finally:
            if not args.keep_temp_csv:
                csv_path.unlink(missing_ok=True)
        if index < len(posts):
            time.sleep(max(0.0, args.between_posts_delay))
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="批量导出监控总表中帖子的评论区 XLSX。")
    parser.add_argument("--platform", choices=["all", "xhs", "dy"], default="all")
    parser.add_argument("--max-posts-per-platform", type=int, default=0, help="调试用：每个平台最多处理多少条；0 表示全部。")
    parser.add_argument("--limit-comments", type=int, default=0, help="调试用：每条帖子最多导出评论数；0 表示全部。")
    parser.add_argument("--request-interval", type=float, default=None, help="评论分页请求间隔；不填则按平台使用保守默认。")
    parser.add_argument("--between-posts-delay", type=float, default=8.0, help="帖子之间等待秒数。")
    parser.add_argument("--include-zero-comments", action="store_true", help="也尝试评论量为 0 或空的帖子。")
    parser.add_argument("--note-id", help="只导出指定笔记ID，多个用英文逗号分隔。")
    parser.add_argument("--overwrite", action="store_true", help="覆盖已存在的同名 XLSX。")
    parser.add_argument("--reset-output", action="store_true", help="先删除所选平台 Comment_Data/{xhs,dy} 下旧 XLSX，再重新导出。")
    parser.add_argument("--dry-run", action="store_true", help="只打印将要导出的文件，不访问平台。")
    parser.add_argument("--headed", action="store_true", help="显示浏览器窗口，便于登录/验证。")
    parser.add_argument("--no-sub-comments", action="store_true", help="只导出一级评论。")
    parser.add_argument("--keep-temp-csv", action="store_true", help="保留中间 CSV。")
    args = parser.parse_args()

    platforms = ["xhs", "dy"] if args.platform == "all" else [args.platform]
    results = [export_platform(platform, args) for platform in platforms]
    log_path = COMMENT_ROOT / f"comment_export_log_{time.strftime('%Y%m%d_%H%M%S')}.json"
    payload = {
        "generatedAt": time.strftime("%Y-%m-%d %H:%M:%S"),
        "dryRun": args.dry_run,
        "results": results,
    }
    log_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({**payload, "log": str(log_path)}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
